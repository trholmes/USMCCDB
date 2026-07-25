"""API tests against a real PostgreSQL (needs TEST_DATABASE_URL, e.g. the
compose db). Skipped automatically when no test database is configured.

Run inside the stack:
    docker compose exec backend sh -c 'TEST_DATABASE_URL=$DATABASE_URL pytest -q'
"""

import os
from datetime import date

import pytest

TEST_DB = os.environ.get("TEST_DATABASE_URL")

pytestmark = pytest.mark.skipif(not TEST_DB, reason="TEST_DATABASE_URL not set")

if TEST_DB:
    os.environ["DATABASE_URL"] = TEST_DB
    os.environ["SECRET_KEY"] = "test-secret"
    os.environ["BOOTSTRAP_ADMIN_USERNAME"] = "testadmin"
    os.environ["BOOTSTRAP_ADMIN_PASSWORD"] = "testadmin-pw"
    # Enable email composition; tests monkeypatch delivery, and any
    # unpatched send fails fast against the invalid host and is logged.
    os.environ["SMTP_HOST"] = "smtp.test.invalid"
    os.environ["CONTACT_EMAIL"] = "office@example.edu"
    # Enable the ORCID endpoints; tests monkeypatch the code exchange.
    os.environ["ORCID_CLIENT_ID"] = "test-client"
    os.environ["ORCID_CLIENT_SECRET"] = "test-secret"
    # Every request here comes from the same TestClient address; the
    # dedicated rate-limit tests dial these back down on the live limiters.
    os.environ["LOGIN_RATE_LIMIT"] = "100000"
    os.environ["REGISTRATION_RATE_LIMIT"] = "100000"

    from fastapi.testclient import TestClient

    from app.db import Base, engine
    from app.main import app


@pytest.fixture(scope="module")
def client():
    from sqlalchemy import text

    with engine.connect() as conn:
        conn.execute(text("CREATE EXTENSION IF NOT EXISTS btree_gist"))
        conn.commit()
    Base.metadata.drop_all(engine)
    Base.metadata.create_all(engine)
    with TestClient(app) as c:  # lifespan runs → bootstrap admin created
        yield c
    Base.metadata.drop_all(engine)


@pytest.fixture(scope="module")
def admin(client):
    resp = client.post(
        "/api/v1/auth/login", json={"username": "testadmin", "password": "testadmin-pw"}
    )
    assert resp.status_code == 200, resp.text
    return client  # cookie persists on the client


def test_health(client):
    assert client.get("/api/v1/health").json() == {"status": "ok"}


def test_startup_refuses_placeholder_secrets(monkeypatch):
    """cp .env.example .env must not yield a bootable app with known secrets
    (issue #59): lifespan refuses placeholder/empty SECRET_KEY and the
    'change-me' bootstrap admin password."""
    from app.config import get_settings

    try:
        for bad in ("change-me-openssl-rand-hex-32", "dev-only-change-me", "  "):
            monkeypatch.setenv("SECRET_KEY", bad)
            get_settings.cache_clear()
            with pytest.raises(RuntimeError, match="SECRET_KEY"):
                with TestClient(app):
                    pass

        monkeypatch.setenv("SECRET_KEY", "test-secret")
        monkeypatch.setenv("BOOTSTRAP_ADMIN_PASSWORD", "change-me")
        get_settings.cache_clear()
        with pytest.raises(RuntimeError, match="BOOTSTRAP_ADMIN_PASSWORD"):
            with TestClient(app):
                pass
    finally:
        # monkeypatch restores the env after this; drop the poisoned cache so
        # later tests re-read the real test settings.
        get_settings.cache_clear()


def test_me_requires_auth(client):
    fresh = TestClient(app)
    assert fresh.get("/api/v1/auth/me").status_code == 401


def test_register_and_approve_flow(admin):
    # Public registration.
    resp = admin.post(
        "/api/v1/people/register",
        json={
            "given_name": "Priya",
            "middle_name": "R.",
            "family_name": "Kumar",
            "email": "priya@example.edu",
            "career_stage": "postdoc",
            "institution_name": "Test University",
            "institution_is_us": True,
            "is_voting": True,
        },
    )
    assert resp.status_code == 201, resp.text
    pid = resp.json()["id"]
    assert resp.json()["status"] == "pending"

    # Office approves.
    resp = admin.post(f"/api/v1/people/{pid}/status", json={"status": "active"})
    assert resp.status_code == 200
    assert resp.json()["status"] == "active"

    # Audit trail recorded both transitions.
    events = admin.get(f"/api/v1/people/{pid}/events").json()
    assert [e["to_status"] for e in events] == ["pending", "active"]

    # The initial affiliation records the stage given at registration.
    person = admin.get(f"/api/v1/people/{pid}").json()
    assert person["affiliations"][0]["career_stage"] == "postdoc"
    assert person["middle_name"] == "R."


def test_author_list_generation(admin):
    # Institution with a formal address.
    inst = admin.post(
        "/api/v1/institutions",
        json={"name": "Univ One", "short_name": "U1", "latex_address": "Univ One, TN, USA"},
    ).json()

    # Two people, alphabetically tricky (accent should not matter).
    p1 = admin.post(
        "/api/v1/people/register",
        json={"given_name": "Zoe", "family_name": "Ábel", "email": "zoe@example.edu"},
    ).json()
    p2 = admin.post(
        "/api/v1/people/register",
        json={"given_name": "Al", "family_name": "Baker", "email": "al@example.edu"},
    ).json()

    for pid in (p1["id"], p2["id"]):
        admin.post(f"/api/v1/people/{pid}/status", json={"status": "active"})
        r = admin.post(
            f"/api/v1/people/{pid}/affiliations",
            json={"institution_id": inst["id"], "is_primary": True, "start_date": "2025-01-01"},
        )
        assert r.status_code == 201, r.text
        r = admin.post(
            f"/api/v1/people/{pid}/author-periods", json={"start_date": "2025-06-01"}
        )
        assert r.status_code == 201, r.text

    # Overlapping author period must be rejected.
    r = admin.post(
        f"/api/v1/people/{p1['id']}/author-periods", json={"start_date": "2026-01-01"}
    )
    assert r.status_code == 409

    # Preview: Ábel sorts before Baker despite the accent.
    snap = admin.post(
        "/api/v1/author-lists/preview", json={"cutoff_date": "2026-07-01"}
    ).json()
    names = [a["family_name"] for a in snap["authors"]]
    assert names.index("Ábel") < names.index("Baker")

    # Cutoff before the periods start → nobody.
    early = admin.post(
        "/api/v1/author-lists/preview", json={"cutoff_date": "2025-01-15"}
    ).json()
    assert early["authors"] == []

    # Full flow via a publication + exports.
    pub = admin.post("/api/v1/publications", json={"title": "First USMCC Paper"}).json()
    alist = admin.post(
        f"/api/v1/publications/{pub['id']}/author-list",
        json={"cutoff_date": "2026-07-01"},
    ).json()
    assert len(alist["snapshot"]["authors"]) >= 2

    tex = admin.get(f"/api/v1/author-lists/{alist['id']}/export?format=tex")
    assert tex.status_code == 200
    assert "\\author[" in tex.text
    xml = admin.get(f"/api/v1/author-lists/{alist['id']}/export?format=xml")
    assert "<collaborationauthorlist" in xml.text


def test_author_list_dedup_and_missing_affiliation_warning(admin):
    inst = admin.post(
        "/api/v1/institutions",
        json={"name": "Dup University", "latex_address": "Dup University, USA"},
    ).json()

    # Two overlapping affiliation rows at the same institution (possible via
    # office add_affiliation or importer re-runs) must yield one id, not two.
    dup = admin.post(
        "/api/v1/people/register",
        json={"given_name": "Dee", "family_name": "Dupe", "email": "dee.dupe@example.edu"},
    ).json()
    admin.post(f"/api/v1/people/{dup['id']}/status", json={"status": "active"})
    r = admin.post(
        f"/api/v1/people/{dup['id']}/affiliations",
        json={"institution_id": inst["id"], "is_primary": True, "start_date": "2025-01-01"},
    )
    assert r.status_code == 201, r.text
    r = admin.post(
        f"/api/v1/people/{dup['id']}/affiliations",
        json={
            "institution_id": inst["id"],
            "is_primary": False,
            "start_date": "2025-06-01",
            "end_date": "2027-01-01",
        },
    )
    assert r.status_code == 201, r.text
    admin.post(f"/api/v1/people/{dup['id']}/author-periods", json={"start_date": "2025-01-01"})

    # An active author period but no affiliation on the cutoff date.
    gap = admin.post(
        "/api/v1/people/register",
        json={"given_name": "Gale", "family_name": "Gapp", "email": "gale.gapp@example.edu"},
    ).json()
    admin.post(f"/api/v1/people/{gap['id']}/status", json={"status": "active"})
    admin.post(f"/api/v1/people/{gap['id']}/author-periods", json={"start_date": "2025-01-01"})

    snap = admin.post(
        "/api/v1/author-lists/preview", json={"cutoff_date": "2026-06-01"}
    ).json()

    dup_row = next(a for a in snap["authors"] if a["person_id"] == dup["id"])
    assert dup_row["institution_ids"] == [inst["id"]]

    gap_row = next(a for a in snap["authors"] if a["person_id"] == gap["id"])
    assert gap_row["institution_ids"] == []
    assert any("Gale Gapp" in w for w in snap["warnings"])
    assert not any("Dee Dupe" in w for w in snap["warnings"])


def test_member_cannot_do_office_things(admin):
    # Create a plain member account.
    r = admin.post(
        "/api/v1/auth/users",
        json={"username": "plainmember", "password": "member-pw-123", "role": "member"},
    )
    assert r.status_code == 201, r.text

    member = TestClient(app)
    assert (
        member.post(
            "/api/v1/auth/login", json={"username": "plainmember", "password": "member-pw-123"}
        ).status_code
        == 200
    )
    # Directory is visible…
    assert member.get("/api/v1/people").status_code == 200
    # …but office/admin actions are forbidden.
    somebody = member.get("/api/v1/people").json()[0]["id"]
    assert member.post(f"/api/v1/people/{somebody}/status", json={"status": "inactive"}).status_code == 403
    assert member.post("/api/v1/institutions", json={"name": "Nope U"}).status_code == 403
    assert member.get("/api/v1/auth/users").status_code == 403


def _linked_member(admin, *, given, family, email, career_stage="postdoc"):
    """Create an active person + a member account linked to them, and return a
    logged-in TestClient plus the person id."""
    person = admin.post(
        "/api/v1/people/register",
        json={
            "given_name": given,
            "family_name": family,
            "email": email,
            "career_stage": career_stage,
        },
    ).json()
    admin.post(f"/api/v1/people/{person['id']}/status", json={"status": "active"})
    uname = email.split("@")[0]
    r = admin.post(
        "/api/v1/auth/users",
        json={
            "username": uname,
            "password": "member-pw-123",
            "role": "member",
            "person_id": person["id"],
        },
    )
    assert r.status_code == 201, r.text
    c = TestClient(app)
    assert c.post(
        "/api/v1/auth/login", json={"username": uname, "password": "member-pw-123"}
    ).status_code == 200
    return c, person["id"]


def test_member_self_status_change_with_effective_date(admin):
    member, pid = _linked_member(
        admin, given="Sam", family="Self", email="sam.self@example.edu"
    )

    # A member may step back as of a date they enter; history records it.
    r = member.post(
        f"/api/v1/people/{pid}/status",
        json={"status": "inactive", "effective_date": "2026-03-15"},
    )
    assert r.status_code == 200, r.text
    assert r.json()["status"] == "inactive"

    events = member.get(f"/api/v1/people/{pid}/events").json()
    last = events[-1]
    assert last["to_status"] == "inactive"
    assert last["effective_date"] == "2026-03-15"

    # Members cannot self-assign moderation states.
    assert member.post(f"/api/v1/people/{pid}/status", json={"status": "rejected"}).status_code == 403
    assert member.post(f"/api/v1/people/{pid}/status", json={"status": "pending"}).status_code == 403


def test_member_self_institution_move_keeps_history(admin):
    member, pid = _linked_member(
        admin, given="Ida", family="Inst", email="ida.inst@example.edu"
    )
    a = admin.post("/api/v1/institutions", json={"name": "Alpha University"}).json()
    b = admin.post("/api/v1/institutions", json={"name": "Beta Institute"}).json()

    # Start at Alpha.
    assert member.post(
        f"/api/v1/people/{pid}/institution",
        json={"institution_id": a["id"], "start_date": "2025-01-01"},
    ).status_code == 201

    # Move to Beta as of a later date — history is preserved.
    assert member.post(
        f"/api/v1/people/{pid}/institution",
        json={"institution_id": b["id"], "start_date": "2026-06-01"},
    ).status_code == 201

    person = member.get(f"/api/v1/people/{pid}").json()
    affils = sorted(person["affiliations"], key=lambda x: x["start_date"])
    assert [x["institution"]["name"] for x in affils] == ["Alpha University", "Beta Institute"]
    # The old affiliation ends ON the move date (issue #67) — ranges are
    # inclusive on both ends, so the person carries both affiliations on the
    # transition day.
    assert affils[0]["end_date"] == "2026-06-01"
    assert affils[1]["end_date"] is None  # new one is open

    # An author list cut exactly on the move date shows both institutions;
    # cut the day after, only the new one.
    for inst_id in (a["id"], b["id"]):
        admin.patch(f"/api/v1/institutions/{inst_id}", json={"latex_address": "addr"})
    admin.post(f"/api/v1/people/{pid}/author-periods", json={"start_date": "2025-01-01"})
    snap = admin.post(
        "/api/v1/author-lists/preview", json={"cutoff_date": "2026-06-01"}
    ).json()
    row = next(x for x in snap["authors"] if x["person_id"] == pid)
    assert row["institution_ids"] == [a["id"], b["id"]]
    snap = admin.post(
        "/api/v1/author-lists/preview", json={"cutoff_date": "2026-06-02"}
    ).json()
    row = next(x for x in snap["authors"] if x["person_id"] == pid)
    assert row["institution_ids"] == [b["id"]]

    # Backdating before the current start is rejected; future dates are
    # rejected; no cross-profile moves.
    assert member.post(
        f"/api/v1/people/{pid}/institution",
        json={"institution_id": a["id"], "start_date": "2025-01-01"},
    ).status_code == 400
    assert member.post(
        f"/api/v1/people/{pid}/institution",
        json={"institution_id": a["id"], "start_date": "2199-01-01"},
    ).status_code == 422
    other = admin.get("/api/v1/people").json()[0]["id"]
    if other != pid:
        assert member.post(
            f"/api/v1/people/{other}/institution",
            json={"institution_id": a["id"], "start_date": "2026-01-01"},
        ).status_code == 403

    # A same-day move is a correction: the superseded row is dropped rather
    # than left as a one-day affiliation.
    assert member.post(
        f"/api/v1/people/{pid}/institution",
        json={"institution_id": a["id"], "start_date": "2026-06-01"},
    ).status_code == 201
    person = member.get(f"/api/v1/people/{pid}").json()
    open_affils = [x for x in person["affiliations"] if x["end_date"] is None]
    assert len(open_affils) == 1 and open_affils[0]["institution"]["id"] == a["id"]
    assert all(x["institution"]["id"] != b["id"] for x in person["affiliations"])


def test_institution_history_records_career_stage(admin):
    member, pid = _linked_member(
        admin, given="Cara", family="Stage", email="cara.stage@example.edu", career_stage="grad"
    )
    a = admin.post("/api/v1/institutions", json={"name": "Gamma College"}).json()
    b = admin.post("/api/v1/institutions", json={"name": "Delta Lab"}).json()

    # A move without a stage stamps the person's current stage.
    assert member.post(
        f"/api/v1/people/{pid}/institution",
        json={"institution_id": a["id"], "start_date": "2025-01-01"},
    ).status_code == 201
    person = member.get(f"/api/v1/people/{pid}").json()
    assert person["affiliations"][0]["career_stage"] == "grad"

    # A move with a new stage updates the profile and the new affiliation,
    # while the closed affiliation keeps the old stage.
    assert member.post(
        f"/api/v1/people/{pid}/institution",
        json={"institution_id": b["id"], "start_date": "2026-06-01", "career_stage": "postdoc"},
    ).status_code == 201
    person = member.get(f"/api/v1/people/{pid}").json()
    assert person["career_stage"] == "postdoc"
    affils = sorted(person["affiliations"], key=lambda x: x["start_date"])
    assert [x["career_stage"] for x in affils] == ["grad", "postdoc"]

    # Editing the stage on the profile keeps the open affiliation in step.
    assert member.patch(f"/api/v1/people/{pid}", json={"career_stage": "staff"}).status_code == 200
    person = member.get(f"/api/v1/people/{pid}").json()
    open_affil = next(x for x in person["affiliations"] if x["end_date"] is None)
    assert open_affil["career_stage"] == "staff"

    # A voting member cannot become a student via an institution move (same
    # rule as a profile edit); nothing about the move is applied.
    assert member.patch(f"/api/v1/people/{pid}", json={"is_voting": True}).status_code == 200
    r = member.post(
        f"/api/v1/people/{pid}/institution",
        json={"institution_id": a["id"], "start_date": "2026-07-01", "career_stage": "grad"},
    )
    assert r.status_code == 422
    person = member.get(f"/api/v1/people/{pid}").json()
    assert person["career_stage"] == "staff"
    open_affil = next(x for x in person["affiliations"] if x["end_date"] is None)
    assert open_affil["institution"]["id"] == b["id"]


def test_institution_people_count(admin):
    a = admin.post("/api/v1/institutions", json={"name": "Count University"}).json()
    b = admin.post("/api/v1/institutions", json={"name": "Empty Tech"}).json()

    pids = []
    for name in ("One", "Two", "Three"):
        p = admin.post(
            "/api/v1/people/register",
            json={
                "given_name": name,
                "family_name": "Counter",
                "email": f"{name.lower()}.counter@example.edu",
            },
        ).json()
        pids.append(p["id"])

    # Two people currently at Count University…
    for pid in pids[:2]:
        assert admin.post(
            f"/api/v1/people/{pid}/affiliations",
            json={"institution_id": a["id"], "start_date": "2025-01-01"},
        ).status_code == 201
    # …and one whose affiliation already ended — past members don't count.
    assert admin.post(
        f"/api/v1/people/{pids[2]}/affiliations",
        json={"institution_id": a["id"], "start_date": "2024-01-01", "end_date": "2024-12-31"},
    ).status_code == 201

    counts = {i["name"]: i["people_count"] for i in admin.get("/api/v1/institutions").json()}
    assert counts["Count University"] == 2
    assert counts["Empty Tech"] == 0
    assert admin.get(f"/api/v1/institutions/{a['id']}").json()["people_count"] == 2


def test_member_voting_eligibility(admin):
    # A grad student cannot self-assign voting membership.
    grad, gid = _linked_member(
        admin, given="Gale", family="Grad", email="gale.grad@example.edu", career_stage="grad"
    )
    assert grad.patch(f"/api/v1/people/{gid}", json={"is_voting": True}).status_code == 422
    # …and neither can the office (the invariant binds every actor).
    assert admin.patch(f"/api/v1/people/{gid}", json={"is_voting": True}).status_code == 422

    # An active postdoc at a US institution can.
    pd, pdid = _linked_member(
        admin, given="Paz", family="Postdoc", email="paz.pd@example.edu", career_stage="postdoc"
    )
    inst = admin.post("/api/v1/institutions", json={"name": "Voting University"}).json()
    assert pd.post(
        f"/api/v1/people/{pdid}/institution",
        json={"institution_id": inst["id"], "start_date": "2025-01-01"},
    ).status_code == 201
    assert pd.patch(f"/api/v1/people/{pdid}", json={"is_voting": True}).status_code == 200
    assert pd.get(f"/api/v1/people/{pdid}").json()["is_voting"] is True

    # Becoming a student again while voting is rejected (invariant enforced).
    assert pd.patch(f"/api/v1/people/{pdid}", json={"career_stage": "grad"}).status_code == 422

    # Explicit nulls on non-nullable fields are a 422, not a 500.
    assert pd.patch(f"/api/v1/people/{pdid}", json={"is_voting": None}).status_code == 422
    assert pd.patch(f"/api/v1/people/{pdid}", json={"career_stage": None}).status_code == 422

    # Stepping back to inactive clears voting automatically.
    assert pd.post(f"/api/v1/people/{pdid}/status", json={"status": "inactive"}).status_code == 200
    assert pd.get(f"/api/v1/people/{pdid}").json()["is_voting"] is False

    # Even with a (hypothetically) stale voting flag, unrelated self-edits
    # must not be blocked by the eligibility check while inactive.
    r = pd.patch(f"/api/v1/people/{pdid}", json={"expertise": "detector R&D"})
    assert r.status_code == 200, r.text


def test_voting_requires_us_institution(admin):
    member, pid = _linked_member(
        admin, given="Uma", family="Usonly", email="uma.usonly@example.edu", career_stage="faculty"
    )
    us = admin.post("/api/v1/institutions", json={"name": "Stateside University"}).json()
    assert us["is_us"] is True  # US is the default
    abroad = admin.post(
        "/api/v1/institutions",
        json={"name": "Overseas Institute", "country": "Switzerland", "is_us": False},
    ).json()

    # Without a current primary affiliation there is no US institution to
    # qualify through.
    assert member.patch(f"/api/v1/people/{pid}", json={"is_voting": True}).status_code == 422

    # At a non-US institution, neither the member nor the office can set the
    # flag (the invariant binds every actor).
    assert member.post(
        f"/api/v1/people/{pid}/institution",
        json={"institution_id": abroad["id"], "start_date": "2025-01-01"},
    ).status_code == 201
    assert member.patch(f"/api/v1/people/{pid}", json={"is_voting": True}).status_code == 422
    assert admin.patch(f"/api/v1/people/{pid}", json={"is_voting": True}).status_code == 422

    # Moving to a US institution unlocks it.
    assert member.post(
        f"/api/v1/people/{pid}/institution",
        json={"institution_id": us["id"], "start_date": "2025-06-01"},
    ).status_code == 201
    assert member.patch(f"/api/v1/people/{pid}", json={"is_voting": True}).status_code == 200

    # A voting member cannot move to a non-US institution while keeping the
    # flag (same rule as becoming a student); nothing about the move applies.
    r = member.post(
        f"/api/v1/people/{pid}/institution",
        json={"institution_id": abroad["id"], "start_date": "2026-01-01"},
    )
    assert r.status_code == 422
    person = member.get(f"/api/v1/people/{pid}").json()
    open_affil = next(x for x in person["affiliations"] if x["end_date"] is None)
    assert open_affil["institution"]["id"] == us["id"]
    assert person["is_voting"] is True

    # The office affiliation route enforces the same rule for open primaries.
    assert admin.post(
        f"/api/v1/people/{pid}/affiliations",
        json={"institution_id": abroad["id"], "is_primary": True, "start_date": "2026-01-01"},
    ).status_code == 422
    # …but recording closed history at a non-US institution is fine.
    assert admin.post(
        f"/api/v1/people/{pid}/affiliations",
        json={
            "institution_id": abroad["id"],
            "is_primary": False,
            "start_date": "2020-01-01",
            "end_date": "2020-12-31",
        },
    ).status_code == 201

    # Reclassifying the institution as non-US clears the voting flag of the
    # people currently there (like a deactivation does).
    assert admin.patch(f"/api/v1/institutions/{us['id']}", json={"is_us": False}).status_code == 200
    assert member.get(f"/api/v1/people/{pid}").json()["is_voting"] is False
    # …and flipping it back does not silently restore the flag.
    assert admin.patch(f"/api/v1/institutions/{us['id']}", json={"is_us": True}).status_code == 200
    assert member.get(f"/api/v1/people/{pid}").json()["is_voting"] is False


def test_registration_voting_eligibility_enforced(admin):
    us = admin.post("/api/v1/institutions", json={"name": "Register US University"}).json()
    abroad = admin.post(
        "/api/v1/institutions",
        json={"name": "Register Overseas Institute", "country": "France", "is_us": False},
    ).json()
    base = {
        "given_name": "Vic",
        "family_name": "Registrant",
        "email": "vic.registrant@example.edu",
        "is_voting": True,
    }

    # A student cannot register as a voting member…
    r = admin.post(
        "/api/v1/people/register",
        json={**base, "career_stage": "grad", "institution_id": us["id"]},
    )
    assert r.status_code == 422
    # …nor can anyone without an institution to qualify through…
    assert admin.post(
        "/api/v1/people/register", json={**base, "career_stage": "postdoc"}
    ).status_code == 422
    # …nor through a non-US institution.
    assert admin.post(
        "/api/v1/people/register",
        json={**base, "career_stage": "postdoc", "institution_id": abroad["id"]},
    ).status_code == 422
    # The rejected attempts created no person record.
    assert all(
        p["email"] != "vic.registrant@example.edu" for p in admin.get("/api/v1/people").json()
    )

    # An eligible registration goes through, and approval keeps the flag.
    r = admin.post(
        "/api/v1/people/register",
        json={**base, "career_stage": "postdoc", "institution_id": us["id"]},
    )
    assert r.status_code == 201, r.text
    pid = r.json()["id"]
    assert r.json()["is_voting"] is True
    assert admin.post(f"/api/v1/people/{pid}/status", json={"status": "active"}).status_code == 200
    assert admin.get(f"/api/v1/people/{pid}").json()["is_voting"] is True


def test_activation_revalidates_voting(admin):
    # A registration naming a new free-text institution is taken at its word…
    r = admin.post(
        "/api/v1/people/register",
        json={
            "given_name": "Renata",
            "family_name": "Recheck",
            "email": "renata.recheck@example.edu",
            "career_stage": "postdoc",
            "institution_name": "Recheck Institute",
            "institution_is_us": True,
            "is_voting": True,
        },
    )
    assert r.status_code == 201, r.text
    pid = r.json()["id"]

    # …but the office review finds it bogus and removes the affiliation
    # before approving, so with no US institution left to qualify through the
    # approval transition drops the requested voting flag instead of granting
    # it unchecked.
    affil_id = admin.get(f"/api/v1/people/{pid}").json()["affiliations"][0]["id"]
    assert admin.delete(f"/api/v1/people/{pid}/affiliations/{affil_id}").status_code == 204
    assert admin.post(f"/api/v1/people/{pid}/status", json={"status": "active"}).status_code == 200
    person = admin.get(f"/api/v1/people/{pid}").json()
    assert person["status"] == "active"
    assert person["is_voting"] is False


def test_affiliation_edit_and_delete_revalidate_voting(admin):
    # Office affiliation PATCH/DELETE must enforce the voting invariant like
    # every other path that touches the involved fields (issue #57).
    us = admin.post("/api/v1/institutions", json={"name": "Affil Edit University"}).json()
    r = admin.post(
        "/api/v1/people/register",
        json={
            "given_name": "Vera",
            "family_name": "Votingedit",
            "email": "vera.votingedit@example.edu",
            "career_stage": "postdoc",
            "institution_id": us["id"],
            "is_voting": True,
        },
    )
    assert r.status_code == 201, r.text
    pid = r.json()["id"]
    assert admin.post(f"/api/v1/people/{pid}/status", json={"status": "active"}).status_code == 200
    assert admin.get(f"/api/v1/people/{pid}").json()["is_voting"] is True
    affil = admin.get(f"/api/v1/people/{pid}").json()["affiliations"][0]
    affil_id = affil["id"]

    # Recording a departure by closing the open primary leaves no current US
    # affiliation — the voting flag is cleared, not left stale. (Closing on
    # the start date keeps the range valid: ranges are inclusive, and the
    # affiliation began today.)
    r = admin.patch(
        f"/api/v1/people/{pid}/affiliations/{affil_id}", json={"end_date": affil["start_date"]}
    )
    assert r.status_code == 200, r.text
    assert admin.get(f"/api/v1/people/{pid}").json()["is_voting"] is False

    # Reopening the affiliation does not silently restore the flag, but the
    # member becomes eligible again and the office can re-grant it.
    assert admin.patch(
        f"/api/v1/people/{pid}/affiliations/{affil_id}", json={"end_date": None}
    ).status_code == 200
    assert admin.get(f"/api/v1/people/{pid}").json()["is_voting"] is False
    assert admin.patch(f"/api/v1/people/{pid}", json={"is_voting": True}).status_code == 200

    # Deleting the open primary outright clears the flag the same way.
    assert admin.delete(f"/api/v1/people/{pid}/affiliations/{affil_id}").status_code == 204
    assert admin.get(f"/api/v1/people/{pid}").json()["is_voting"] is False


def test_affiliation_edit_second_open_primary_is_409(admin):
    # Editing a second affiliation into an open primary trips
    # uq_one_open_primary_affiliation — a clean 409, not a 500 (issue #57).
    us = admin.post("/api/v1/institutions", json={"name": "Affil Conflict University"}).json()
    other = admin.post("/api/v1/institutions", json={"name": "Affil Conflict Institute"}).json()
    r = admin.post(
        "/api/v1/people/register",
        json={
            "given_name": "Colin",
            "family_name": "Conflictson",
            "email": "colin.conflictson@example.edu",
            "career_stage": "postdoc",
            "institution_id": us["id"],
        },
    )
    assert r.status_code == 201, r.text
    pid = r.json()["id"]
    assert admin.post(f"/api/v1/people/{pid}/status", json={"status": "active"}).status_code == 200
    r = admin.post(
        f"/api/v1/people/{pid}/affiliations",
        json={
            "institution_id": other["id"],
            "is_primary": False,
            "start_date": "2020-01-01",
            "end_date": "2020-12-31",
        },
    )
    assert r.status_code == 201, r.text
    second_id = r.json()["id"]

    r = admin.patch(
        f"/api/v1/people/{pid}/affiliations/{second_id}",
        json={"is_primary": True, "end_date": None},
    )
    assert r.status_code == 409, r.text
    # The rejected edit was rolled back entirely.
    person = admin.get(f"/api/v1/people/{pid}").json()
    second = next(a for a in person["affiliations"] if a["id"] == second_id)
    assert second["is_primary"] is False
    assert second["end_date"] == "2020-12-31"


def test_inverted_date_ranges_rejected(admin):
    # An affiliation or author period with end_date < start_date matches no
    # cutoff date — the person silently drops off every generated author
    # list — so inverted ranges must be rejected up front (issue #61).
    inst = admin.post("/api/v1/institutions", json={"name": "Inverted Range University"}).json()
    r = admin.post(
        "/api/v1/people/register",
        json={
            "given_name": "Ivy",
            "family_name": "Inverted",
            "email": "ivy.inverted@example.edu",
            "career_stage": "postdoc",
            "institution_id": inst["id"],
        },
    )
    assert r.status_code == 201, r.text
    pid = r.json()["id"]
    assert admin.post(f"/api/v1/people/{pid}/status", json={"status": "active"}).status_code == 200

    # Creating an inverted affiliation is rejected by the schema...
    r = admin.post(
        f"/api/v1/people/{pid}/affiliations",
        json={"institution_id": inst["id"], "start_date": "2024-01-01", "end_date": "2023-01-01"},
    )
    assert r.status_code == 422, r.text

    # ...and a partial edit that inverts the stored range is caught too.
    affil_id = admin.get(f"/api/v1/people/{pid}").json()["affiliations"][0]["id"]
    r = admin.patch(
        f"/api/v1/people/{pid}/affiliations/{affil_id}", json={"end_date": "1990-01-01"}
    )
    assert r.status_code == 422, r.text
    r = admin.patch(
        f"/api/v1/people/{pid}/affiliations/{affil_id}",
        json={"start_date": "2024-01-01", "end_date": "2023-01-01"},
    )
    assert r.status_code == 422, r.text
    # The rejected edits left the affiliation open.
    person = admin.get(f"/api/v1/people/{pid}").json()
    assert next(a for a in person["affiliations"] if a["id"] == affil_id)["end_date"] is None

    # Author periods: an inverted create is a 422, not a misleading
    # "overlapping" 409 from the bare daterange rejection.
    r = admin.post(
        f"/api/v1/people/{pid}/author-periods",
        json={"start_date": "2024-01-01", "end_date": "2023-01-01"},
    )
    assert r.status_code == 422, r.text

    # A valid period cannot be inverted by a partial edit either...
    r = admin.post(
        f"/api/v1/people/{pid}/author-periods",
        json={"start_date": "2024-01-01", "end_date": "2024-12-31"},
    )
    assert r.status_code == 201, r.text
    period_id = r.json()["id"]
    r = admin.patch(
        f"/api/v1/people/{pid}/author-periods/{period_id}", json={"end_date": "2023-06-01"}
    )
    assert r.status_code == 422, r.text
    periods = admin.get(f"/api/v1/people/{pid}/author-periods").json()
    assert next(p for p in periods if p["id"] == period_id)["end_date"] == "2024-12-31"

    # ...and a genuine overlap still maps to 409.
    r = admin.post(f"/api/v1/people/{pid}/author-periods", json={"start_date": "2024-06-01"})
    assert r.status_code == 409, r.text


def test_free_text_institution_requires_us_declaration(admin):
    base = {
        "given_name": "Nina",
        "family_name": "Newinst",
        "email": "nina.newinst@example.edu",
        "career_stage": "postdoc",
    }
    # Naming a new institution without declaring it US / non-US is rejected —
    # is_us gates voting eligibility, so the backend never guesses (issue #56).
    r = admin.post(
        "/api/v1/people/register",
        json={**base, "institution_name": "Somewhere Institute"},
    )
    assert r.status_code == 422

    # A declared non-US institution closes the free-text voting bypass.
    r = admin.post(
        "/api/v1/people/register",
        json={
            **base,
            "institution_name": "University of Elsewhere",
            "institution_is_us": False,
            "is_voting": True,
        },
    )
    assert r.status_code == 422

    # A non-voting registration goes through, and the new (inactive)
    # institution records the declaration.
    r = admin.post(
        "/api/v1/people/register",
        json={
            **base,
            "institution_name": "University of Elsewhere",
            "institution_is_us": False,
        },
    )
    assert r.status_code == 201, r.text
    inst = next(
        i
        for i in admin.get("/api/v1/institutions").json()
        if i["name"] == "University of Elsewhere"
    )
    assert inst["is_us"] is False
    assert inst["is_active"] is False
    assert inst["country"] is None


def test_institution_move_free_text_requires_us_declaration(admin):
    member, pid = _linked_member(
        admin, given="Mo", family="Mover", email="mo.mover@example.edu"
    )
    us = admin.post("/api/v1/institutions", json={"name": "Stateside U"}).json()
    assert member.post(
        f"/api/v1/people/{pid}/institution",
        json={"institution_id": us["id"], "start_date": "2025-01-01"},
    ).status_code == 201
    assert member.patch(f"/api/v1/people/{pid}", json={"is_voting": True}).status_code == 200

    # A free-text move must declare the new institution US / non-US.
    r = member.post(
        f"/api/v1/people/{pid}/institution",
        json={"institution_name": "University of Tokyo", "start_date": "2026-01-01"},
    )
    assert r.status_code == 422

    # A voting member cannot move to a declared non-US institution while
    # keeping the flag (the issue #56 bypass).
    r = member.post(
        f"/api/v1/people/{pid}/institution",
        json={
            "institution_name": "University of Tokyo",
            "institution_is_us": False,
            "start_date": "2026-01-01",
        },
    )
    assert r.status_code == 422
    assert member.get(f"/api/v1/people/{pid}").json()["is_voting"] is True

    # After giving up the flag the move goes through, and the non-US record
    # blocks re-granting voting membership.
    assert member.patch(f"/api/v1/people/{pid}", json={"is_voting": False}).status_code == 200
    r = member.post(
        f"/api/v1/people/{pid}/institution",
        json={
            "institution_name": "University of Tokyo",
            "institution_is_us": False,
            "start_date": "2026-01-01",
        },
    )
    assert r.status_code == 201, r.text
    assert r.json()["institution"]["is_us"] is False
    assert member.patch(f"/api/v1/people/{pid}", json={"is_voting": True}).status_code == 422


def test_research_areas_normalized(admin):
    member, pid = _linked_member(
        admin, given="Ria", family="Areas", email="ria.areas@example.edu"
    )
    # Standard categories are normalized: case, spacing, and duplicates.
    r = member.patch(
        f"/api/v1/people/{pid}",
        json={
            "research_areas": "experimental particle physics,  ACCELERATOR PHYSICS,"
            " Experimental Particle Physics"
        },
    )
    assert r.status_code == 200, r.text
    assert r.json()["research_areas"] == "Experimental Particle Physics, Accelerator Physics"
    # Values outside the standard set are rejected.
    r = member.patch(f"/api/v1/people/{pid}", json={"research_areas": "magnets"})
    assert r.status_code == 422
    # The field can be cleared, and free-form topics live in expertise.
    r = member.patch(
        f"/api/v1/people/{pid}", json={"research_areas": None, "expertise": "muon cooling, MDI"}
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["research_areas"] is None
    assert body["expertise"] == "muon cooling, MDI"


def test_directory_filters(admin):
    member, pid = _linked_member(
        admin, given="Fay", family="Filter", email="fay.filter@example.edu", career_stage="staff"
    )
    other, oid = _linked_member(
        admin, given="Ned", family="Nofilter", email="ned.nofilter@example.edu"
    )
    inst = admin.post("/api/v1/institutions", json={"name": "Filter University"}).json()
    assert member.post(
        f"/api/v1/people/{pid}/institution",
        json={"institution_id": inst["id"], "start_date": "2025-01-01"},
    ).status_code == 201
    assert member.patch(
        f"/api/v1/people/{pid}",
        json={"research_areas": "Accelerator Physics, Other", "is_voting": True},
    ).status_code == 200

    # Research-area filter matches case-insensitively against the canonical
    # names, and the summary rows expose research_areas for the frontend.
    r = member.get("/api/v1/people?research_area=accelerator physics")
    assert r.status_code == 200, r.text
    ids = {p["id"] for p in r.json()}
    assert pid in ids and oid not in ids
    row = next(p for p in r.json() if p["id"] == pid)
    assert row["research_areas"] == "Accelerator Physics, Other"

    # Unknown areas are rejected rather than silently matching nothing.
    assert member.get("/api/v1/people?research_area=magnets").status_code == 422

    # Voting filter, both polarities.
    voting = {p["id"] for p in member.get("/api/v1/people?is_voting=true").json()}
    assert pid in voting and oid not in voting
    nonvoting = {p["id"] for p in member.get("/api/v1/people?is_voting=false").json()}
    assert oid in nonvoting and pid not in nonvoting

    # Filters combine.
    r = member.get("/api/v1/people?research_area=Other&career_stage=staff&is_voting=true")
    assert [p["id"] for p in r.json()] == [pid]


def test_office_status_change_clears_voting(admin):
    _, pid = _linked_member(
        admin, given="Vera", family="Vote", email="vera.vote@example.edu", career_stage="faculty"
    )
    inst = admin.post("/api/v1/institutions", json={"name": "Vote Lab"}).json()
    assert admin.post(
        f"/api/v1/people/{pid}/affiliations",
        json={"institution_id": inst["id"], "is_primary": True, "start_date": "2025-01-01"},
    ).status_code == 201
    assert admin.patch(f"/api/v1/people/{pid}", json={"is_voting": True}).status_code == 200

    # Office deactivation must clear the voting flag too, not just self-service.
    assert admin.post(f"/api/v1/people/{pid}/status", json={"status": "inactive"}).status_code == 200
    assert admin.get(f"/api/v1/people/{pid}").json()["is_voting"] is False


def test_member_cannot_self_reinstate(admin):
    # A pending registration's login gets no API access at all (issue #50):
    # the password sign-in is refused with a clear reason…
    person = admin.post(
        "/api/v1/people/register",
        json={"given_name": "Pat", "family_name": "Pending", "email": "pat.pending@example.edu"},
    ).json()
    r = admin.post(
        "/api/v1/auth/users",
        json={
            "username": "patpending",
            "password": "member-pw-123",
            "role": "member",
            "person_id": person["id"],
        },
    )
    assert r.status_code == 201, r.text
    member = TestClient(app)
    r = member.post(
        "/api/v1/auth/login", json={"username": "patpending", "password": "member-pw-123"}
    )
    assert r.status_code == 403
    assert "awaiting approval" in r.json()["detail"]

    # …so a pending member cannot self-approve either.
    assert member.post(
        f"/api/v1/people/{person['id']}/status", json={"status": "active"}
    ).status_code == 401  # never signed in

    # Once approved, sign-in works and the directory opens up.
    admin.post(f"/api/v1/people/{person['id']}/status", json={"status": "active"})
    assert member.post(
        "/api/v1/auth/login", json={"username": "patpending", "password": "member-pw-123"}
    ).status_code == 200
    assert member.get("/api/v1/people").status_code == 200

    # Rejection cuts off an existing session immediately (the status is read
    # per request, not baked into the token)…
    admin.post(f"/api/v1/people/{person['id']}/status", json={"status": "rejected"})
    r = member.get("/api/v1/people")
    assert r.status_code == 403
    assert "not approved" in r.json()["detail"]
    # …and a rejected member cannot reinstate themselves.
    assert member.post(
        f"/api/v1/people/{person['id']}/status", json={"status": "active"}
    ).status_code == 403

    # Future-dated status changes are rejected outright.
    admin.post(f"/api/v1/people/{person['id']}/status", json={"status": "active"})
    assert member.post(
        f"/api/v1/people/{person['id']}/status",
        json={"status": "inactive", "effective_date": "2199-01-01"},
    ).status_code == 422


def test_membership_notes_hidden_from_members(admin):
    member, pid = _linked_member(
        admin, given="Nia", family="Note", email="nia.note@example.edu"
    )
    admin.post(
        f"/api/v1/people/{pid}/status",
        json={"status": "inactive", "note": "office-internal: verify affiliation"},
    )
    # Office sees the note; the member sees the transition but not the note.
    admin_events = admin.get(f"/api/v1/people/{pid}/events").json()
    assert any(e["note"] == "office-internal: verify affiliation" for e in admin_events)
    member_events = member.get(f"/api/v1/people/{pid}/events").json()
    assert [e["to_status"] for e in member_events] == [e["to_status"] for e in admin_events]
    assert all(e["note"] is None and e["actor_user_id"] is None for e in member_events)


def test_person_notes_hidden_from_members(admin):
    member, pid = _linked_member(
        admin, given="Priya", family="Private", email="priya.private@example.edu"
    )
    r = admin.patch(
        f"/api/v1/people/{pid}", json={"notes": "application dubious — verify employment"}
    )
    assert r.status_code == 200, r.text
    # Office sees the note on the profile; everyone else — including the
    # person themselves — gets it scrubbed (same policy as event notes).
    assert admin.get(f"/api/v1/people/{pid}").json()["notes"] == (
        "application dubious — verify employment"
    )
    assert member.get(f"/api/v1/people/{pid}").json()["notes"] is None
    # The PersonOut returned by a self-edit must not leak it either.
    r = member.patch(f"/api/v1/people/{pid}", json={"preferred_name": "Pri"})
    assert r.status_code == 200, r.text
    assert r.json()["notes"] is None


def test_photo_upload_and_serve(admin, tmp_path_factory):
    os.environ["PHOTOS_DIR"] = str(tmp_path_factory.mktemp("photos"))
    from app.config import get_settings

    get_settings.cache_clear()

    person = admin.get("/api/v1/people").json()[0]
    png = bytes.fromhex(
        "89504e470d0a1a0a0000000d494844520000000100000001080600000"
        "01f15c4890000000d49444154789c6260010000000500010d0a2db400"
        "00000049454e44ae426082"
    )
    r = admin.post(
        f"/api/v1/people/{person['id']}/photo",
        files={"file": ("me.png", png, "image/png")},
    )
    assert r.status_code == 201, r.text
    assert r.json()["photo_file"]

    served = admin.get(f"/api/v1/people/{person['id']}/photo")
    assert served.status_code == 200
    assert served.content == png

    # Wrong content type is rejected.
    bad = admin.post(
        f"/api/v1/people/{person['id']}/photo",
        files={"file": ("evil.svg", b"<svg/>", "image/svg+xml")},
    )
    assert bad.status_code == 422

    assert admin.delete(f"/api/v1/people/{person['id']}/photo").status_code == 204
    assert admin.get(f"/api/v1/people/{person['id']}/photo").status_code == 404


def test_photo_upload_rejects_bad_content(admin, tmp_path_factory, monkeypatch):
    os.environ["PHOTOS_DIR"] = str(tmp_path_factory.mktemp("photos"))
    from app.config import get_settings

    get_settings.cache_clear()
    person = admin.get("/api/v1/people").json()[0]
    before = admin.get(f"/api/v1/people/{person['id']}").json()["photo_file"]

    # An allowed Content-Type whose bytes are not that format is rejected.
    fake = admin.post(
        f"/api/v1/people/{person['id']}/photo",
        files={"file": ("fake.png", b"<html>not an image</html>", "image/png")},
    )
    assert fake.status_code == 422
    assert admin.get(f"/api/v1/people/{person['id']}").json()["photo_file"] == before

    # An oversized body is rejected (shrink the limit rather than posting 10 MB).
    from app.routers import people as people_router

    monkeypatch.setattr(people_router, "MAX_PHOTO_BYTES", 1024)
    big = admin.post(
        f"/api/v1/people/{person['id']}/photo",
        files={"file": ("big.png", b"\x89PNG\r\n\x1a\n" + b"\0" * 2048, "image/png")},
    )
    assert big.status_code == 413


def test_speakers_flow(admin):
    event = admin.post("/api/v1/events", json={"name": "Snowmass 2026"}).json()
    talk = admin.post(
        "/api/v1/talks",
        json={
            "title": "Muon Collider Status",
            "event_id": event["id"],
            "talk_type": "plenary",
            "date": "2026-08-01",
            "is_invited": True,
        },
    ).json()

    person = admin.get("/api/v1/people").json()[0]
    nom = admin.post(
        f"/api/v1/talks/{talk['id']}/nominations", json={"person_id": person["id"]}
    )
    assert nom.status_code == 201, nom.text

    # Assigning the nomination sets the talk's speaker.
    r = admin.patch(f"/api/v1/nominations/{nom.json()['id']}", json={"status": "assigned"})
    assert r.status_code == 200
    talk_now = admin.get(f"/api/v1/talks?event_id={event['id']}").json()[0]
    assert talk_now["speaker_person_id"] == person["id"]
    assert talk_now["status"] == "assigned"

    stats = admin.get("/api/v1/stats/talks?by=person").json()
    row = next(s for s in stats if s["key_id"] == person["id"])
    assert row["talks"] >= 1 and row["invited"] >= 1


def test_withdraw_assigned_nomination_resets_talk(admin):
    """Withdrawing an assigned nomination must release the talk instead of
    leaving the withdrawn speaker assigned (issue #64)."""
    member, pid = _linked_member(
        admin, given="Wilma", family="Withdrawer", email="wilma.withdrawer@example.edu"
    )
    event = admin.post("/api/v1/events", json={"name": "Withdrawal Workshop"}).json()
    talk = admin.post(
        "/api/v1/talks",
        json={"title": "Cooling Channel Update", "event_id": event["id"], "date": "2026-10-01"},
    ).json()

    nom = admin.post(f"/api/v1/talks/{talk['id']}/nominations", json={"person_id": pid}).json()
    r = admin.patch(f"/api/v1/nominations/{nom['id']}", json={"status": "assigned"})
    assert r.status_code == 200

    # The assigned speaker withdraws themselves: talk goes back to open with
    # no speaker (no other live nominations remain).
    r = member.patch(f"/api/v1/nominations/{nom['id']}", json={"status": "withdrawn"})
    assert r.status_code == 200, r.text
    talk_now = admin.get(f"/api/v1/talks?event_id={event['id']}").json()[0]
    assert talk_now["speaker_person_id"] is None
    assert talk_now["status"] == "open"
    stats = admin.get("/api/v1/stats/talks?by=person").json()
    assert not any(s["key_id"] == pid for s in stats)

    # With another live nomination the talk drops back to "nominations" instead.
    other = admin.post(
        "/api/v1/people/register",
        json={
            "given_name": "Norm",
            "family_name": "Nominee",
            "email": "norm.nominee@example.edu",
            "career_stage": "postdoc",
        },
    ).json()
    admin.post(f"/api/v1/talks/{talk['id']}/nominations", json={"person_id": other["id"]})
    r = admin.patch(f"/api/v1/nominations/{nom['id']}", json={"status": "assigned"})
    assert r.status_code == 200
    r = admin.patch(f"/api/v1/nominations/{nom['id']}", json={"status": "withdrawn"})
    assert r.status_code == 200
    talk_now = admin.get(f"/api/v1/talks?event_id={event['id']}").json()[0]
    assert talk_now["speaker_person_id"] is None
    assert talk_now["status"] == "nominations"

    # Office demoting the assigned nomination to shortlisted also unassigns,
    # and the still-live nomination keeps the talk in "nominations".
    other_nom = admin.get(f"/api/v1/talks?event_id={event['id']}").json()[0]["nominations"]
    other_nom_id = next(n["id"] for n in other_nom if n["person"]["id"] == other["id"])
    r = admin.patch(f"/api/v1/nominations/{other_nom_id}", json={"status": "assigned"})
    assert r.status_code == 200
    r = admin.patch(f"/api/v1/nominations/{other_nom_id}", json={"status": "shortlisted"})
    assert r.status_code == 200
    talk_now = admin.get(f"/api/v1/talks?event_id={event['id']}").json()[0]
    assert talk_now["speaker_person_id"] is None
    assert talk_now["status"] == "nominations"


def test_dangling_references_are_404(admin):
    # Nonexistent FK targets on talk / publication writes must come back as
    # 404s, not unhandled IntegrityError 500s (issue #61).
    event = admin.post("/api/v1/events", json={"name": "FK Check Workshop"}).json()
    r = admin.post("/api/v1/talks", json={"title": "Ghost speaker", "speaker_person_id": 999999})
    assert r.status_code == 404, r.text
    r = admin.post("/api/v1/talks", json={"title": "Ghost WG", "working_group_id": 999999})
    assert r.status_code == 404, r.text

    talk = admin.post("/api/v1/talks", json={"title": "Real talk", "event_id": event["id"]}).json()
    for patch in (
        {"event_id": 999999},
        {"speaker_person_id": 999999},
        {"working_group_id": 999999},
    ):
        r = admin.patch(f"/api/v1/talks/{talk['id']}", json=patch)
        assert r.status_code == 404, r.text
    # Clearing a reference is still allowed.
    r = admin.patch(f"/api/v1/talks/{talk['id']}", json={"event_id": None})
    assert r.status_code == 200 and r.json()["event_id"] is None

    pub = admin.post("/api/v1/publications", json={"title": "FK Check Paper"}).json()
    r = admin.patch(f"/api/v1/publications/{pub['id']}", json={"working_group_id": 999999})
    assert r.status_code == 404, r.text

    assert admin.delete(f"/api/v1/talks/{talk['id']}").status_code == 204


def test_member_self_service_colloquia(admin):
    """Members can record their own seminars/colloquia — talks with a venue
    but no conference — and manage only the talks they added (issue #33)."""
    member, pid = _linked_member(
        admin, given="Sem", family="Speaker", email="sem.speaker@example.edu"
    )

    r = member.post(
        "/api/v1/talks",
        json={
            "title": "Muon Colliders 101",
            "venue": "MIT physics colloquium",
            "talk_type": "colloquium",
            "date": "2026-03-02",
            "speaker_person_id": pid,
            "status": "given",
            "is_invited": True,
        },
    )
    assert r.status_code == 201, r.text
    talk = r.json()
    assert talk["event_id"] is None
    assert talk["venue"] == "MIT physics colloquium"

    # It shows up in the shared talks list alongside conference talks.
    listed = member.get("/api/v1/talks").json()
    assert any(t["id"] == talk["id"] for t in listed)
    mine = member.get(f"/api/v1/talks?speaker_person_id={pid}").json()
    assert [t["id"] for t in mine] == [talk["id"]]

    # The creator may fix their own entry; others' talks stay office-only.
    r = member.patch(f"/api/v1/talks/{talk['id']}", json={"venue": "MIT LNS colloquium"})
    assert r.status_code == 200 and r.json()["venue"] == "MIT LNS colloquium"
    office_talk = admin.post("/api/v1/talks", json={"title": "Office seminar"}).json()
    assert member.patch(
        f"/api/v1/talks/{office_talk['id']}", json={"title": "hijacked"}
    ).status_code == 403
    assert member.delete(f"/api/v1/talks/{office_talk['id']}").status_code == 403

    # Colloquia count toward the fair-share talk statistics.
    stats = member.get("/api/v1/stats/talks?by=person").json()
    row = next(s for s in stats if s["key_id"] == pid)
    assert row["talks"] == 1 and row["invited"] == 1

    assert member.delete(f"/api/v1/talks/{talk['id']}").status_code == 204
    assert admin.delete(f"/api/v1/talks/{office_talk['id']}").status_code == 204


def test_collab_roles_lifecycle(admin):
    member, pid = _linked_member(
        admin, given="Lea", family="Lead", email="lea.lead@example.edu"
    )

    # Simple roles need no qualifier.
    chair = admin.post(
        "/api/v1/collab-roles",
        json={"person_id": pid, "role": "chair", "start_date": "2025-01-01"},
    )
    assert chair.status_code == 201, chair.text

    # Generic organigram roles require a detail qualifier…
    r = admin.post(
        "/api/v1/collab-roles",
        json={"person_id": pid, "role": "representative", "start_date": "2025-01-01"},
    )
    assert r.status_code == 422
    # …and blank detail does not count.
    r = admin.post(
        "/api/v1/collab-roles",
        json={
            "person_id": pid,
            "role": "coordinator",
            "detail": "   ",
            "start_date": "2025-01-01",
        },
    )
    assert r.status_code == 422

    rep = admin.post(
        "/api/v1/collab-roles",
        json={
            "person_id": pid,
            "role": "representative",
            "detail": "Accelerator",
            "start_date": "2025-01-01",
        },
    )
    assert rep.status_code == 201, rep.text
    assert rep.json()["detail"] == "Accelerator"

    # Scoped roles keep their existing requirements.
    assert admin.post(
        "/api/v1/collab-roles",
        json={"person_id": pid, "role": "convener", "start_date": "2025-01-01"},
    ).status_code == 422

    # Listings resolve the person (for the leadership page).
    roles = admin.get(f"/api/v1/collab-roles?person_id={pid}").json()
    assert {x["role"] for x in roles} == {"chair", "representative"}
    assert all(x["person"]["family_name"] == "Lead" for x in roles)

    # Office can close a term; dates stay editable.
    rep_id = rep.json()["id"]
    r = admin.patch(f"/api/v1/collab-roles/{rep_id}", json={"end_date": "2026-06-30"})
    assert r.status_code == 200 and r.json()["end_date"] == "2026-06-30"
    # …but cannot blank a required qualifier.
    assert admin.patch(
        f"/api/v1/collab-roles/{rep_id}", json={"detail": None}
    ).status_code == 422

    # Members can read but not manage roles.
    assert member.get("/api/v1/collab-roles").status_code == 200
    assert member.post(
        "/api/v1/collab-roles",
        json={"person_id": pid, "role": "chair", "start_date": "2025-01-01"},
    ).status_code == 403
    assert member.patch(
        f"/api/v1/collab-roles/{rep_id}", json={"end_date": "2025-12-31"}
    ).status_code == 403
    assert member.delete(f"/api/v1/collab-roles/{rep_id}").status_code == 403

    assert admin.delete(f"/api/v1/collab-roles/{rep_id}").status_code == 204


def test_admin_contact_role_and_scoped_edits(admin):
    inst = admin.post("/api/v1/institutions", json={"name": "Contact University"}).json()
    other = admin.post("/api/v1/institutions", json={"name": "Elsewhere Institute"}).json()

    contact, contact_pid = _linked_member(
        admin, given="Ada", family="Contact", email="ada.contact@example.edu",
        career_stage="staff",
    )
    member, member_pid = _linked_member(
        admin, given="Mo", family="Member", email="mo.member@example.edu"
    )
    _, away_pid = _linked_member(
        admin, given="Ana", family="Away", email="ana.away@example.edu"
    )
    for pid, inst_id in (
        (contact_pid, inst["id"]),
        (member_pid, inst["id"]),
        (away_pid, other["id"]),
    ):
        r = admin.post(
            f"/api/v1/people/{pid}/affiliations",
            json={"institution_id": inst_id, "is_primary": True, "start_date": "2025-01-01"},
        )
        assert r.status_code == 201, r.text

    # The charter role is institution-scoped: no institution_id → rejected.
    r = admin.post(
        "/api/v1/collab-roles",
        json={"person_id": contact_pid, "role": "admin_contact", "start_date": "2025-01-01"},
    )
    assert r.status_code == 422
    role = admin.post(
        "/api/v1/collab-roles",
        json={
            "person_id": contact_pid,
            "role": "admin_contact",
            "institution_id": inst["id"],
            "start_date": "2025-01-01",
        },
    )
    assert role.status_code == 201, role.text

    # Institution-scoped listing feeds the institution detail page.
    rows = admin.get(
        f"/api/v1/collab-roles?institution_id={inst['id']}&role=admin_contact"
    ).json()
    assert [x["person_id"] for x in rows] == [contact_pid]

    # Members record the charter institutional info on their own profile…
    r = member.patch(
        f"/api/v1/people/{member_pid}",
        json={
            "professional_title": "Research Scientist",
            "department": "Physics",
            "usmcc_percent": 50,
        },
    )
    assert r.status_code == 200, r.text
    assert r.json()["professional_title"] == "Research Scientist"
    assert r.json()["usmcc_percent"] == 50
    # …bounded to a sensible percent range.
    assert member.patch(
        f"/api/v1/people/{member_pid}", json={"usmcc_percent": 150}
    ).status_code == 422

    # The registration form collects the same info.
    applied = admin.post(
        "/api/v1/people/register",
        json={
            "given_name": "Tia",
            "family_name": "Titled",
            "email": "tia.titled@example.edu",
            "professional_title": "Beam Physicist",
            "department": "Accelerator Division",
            "usmcc_percent": 25,
        },
    )
    assert applied.status_code == 201, applied.text
    fetched = admin.get(f"/api/v1/people/{applied.json()['id']}").json()
    assert fetched["professional_title"] == "Beam Physicist"
    assert fetched["department"] == "Accelerator Division"
    assert fetched["usmcc_percent"] == 25

    # The admin contact keeps that info up to date for people currently at
    # their institution…
    r = contact.patch(
        f"/api/v1/people/{member_pid}",
        json={"department": "Physics & Astronomy", "usmcc_percent": 40, "career_stage": "staff"},
    )
    assert r.status_code == 200, r.text
    assert r.json()["department"] == "Physics & Astronomy"
    assert r.json()["career_stage"] == "staff"
    # …but cannot touch anything else…
    assert contact.patch(
        f"/api/v1/people/{member_pid}", json={"email": "hijack@example.edu"}
    ).status_code == 403
    # …and has no reach into other institutions.
    assert contact.patch(
        f"/api/v1/people/{away_pid}", json={"usmcc_percent": 10}
    ).status_code == 403

    # An ended term grants nothing.
    r = admin.patch(
        f"/api/v1/collab-roles/{role.json()['id']}", json={"end_date": "2025-12-31"}
    )
    assert r.status_code == 200
    assert contact.patch(
        f"/api/v1/people/{member_pid}", json={"usmcc_percent": 30}
    ).status_code == 403


def test_member_publication_flow(admin, monkeypatch):
    """Any member can register a publication, attach involved people, build a
    subset author list, and request collaboration review — but only the office
    can move it further or assign reviewers. Workflow steps notify the right
    people by email."""
    import app.services.email as email_mod

    sent = []
    monkeypatch.setattr(email_mod, "_deliver", sent.append)

    editor, editor_pid = _linked_member(
        admin, given="Erin", family="Editor", email="erin.editor@example.edu"
    )
    friend, friend_pid = _linked_member(
        admin, given="Finn", family="Friend", email="finn.friend@example.edu"
    )

    r = editor.post(
        "/api/v1/publications",
        json={"title": "A Democratized Paper", "abstract": "Anyone can start one."},
    )
    assert r.status_code == 201, r.text
    pub = r.json()
    assert pub["status"] == "in_progress"
    # The creator is automatically an editor…
    assert [(pp["person"]["id"], pp["role"]) for pp in pub["people"]] == [(editor_pid, "editor")]

    # …and may edit the record and attach involved people from the directory.
    assert editor.patch(
        f"/api/v1/publications/{pub['id']}", json={"target_journal": "PRD"}
    ).status_code == 200
    r = editor.post(
        f"/api/v1/publications/{pub['id']}/people",
        json={"person_id": friend_pid, "role": "contributor"},
    )
    assert r.status_code == 201, r.text
    # Reviewers stay office-assigned.
    assert editor.post(
        f"/api/v1/publications/{pub['id']}/people",
        json={"person_id": friend_pid, "role": "reviewer"},
    ).status_code == 403

    # An uninvolved member can neither edit nor attach people.
    assert friend.patch(
        f"/api/v1/publications/{pub['id']}", json={"title": "Hijacked"}
    ).status_code == 403
    assert friend.post(
        f"/api/v1/publications/{pub['id']}/people",
        json={"person_id": friend_pid, "role": "editor"},
    ).status_code == 403

    # Author list from just the involved people (no author periods needed).
    r = editor.post(
        f"/api/v1/publications/{pub['id']}/author-list",
        json={"cutoff_date": "2026-07-01", "scope": "involved"},
    )
    assert r.status_code == 201, r.text
    names = sorted(a["family_name"] for a in r.json()["snapshot"]["authors"])
    assert names == ["Editor", "Friend"]

    # Members cannot skip ahead in the workflow…
    assert editor.post(
        f"/api/v1/publications/{pub['id']}/status", json={"status": "published"}
    ).status_code == 403
    assert friend.post(
        f"/api/v1/publications/{pub['id']}/status", json={"status": "collab_review"}
    ).status_code == 403
    # …but an editor may request collaboration review, which emails the office.
    sent.clear()
    r = editor.post(
        f"/api/v1/publications/{pub['id']}/status", json={"status": "collab_review"}
    )
    assert r.status_code == 200, r.text
    assert r.json()["status"] == "collab_review"
    assert [(m["To"], m["Subject"]) for m in sent] == [
        ("office@example.edu", "Collaboration review requested: A Democratized Paper")
    ]
    assert "Erin Editor" in sent[0].get_content()

    # Suggested acknowledgment: generic until the office assigns reviewers.
    ack = editor.get(f"/api/v1/publications/{pub['id']}/acknowledgment").json()
    assert "US Muon Collider Collaboration" in ack["text"]
    assert ack["reviewers"] == []

    reviewer_person = admin.post(
        "/api/v1/people/register",
        json={
            "given_name": "Rae",
            "family_name": "Reviewer",
            "email": "rae.reviewer@example.edu",
            "career_stage": "faculty",
        },
    ).json()
    # Assigning a reviewer emails them…
    sent.clear()
    assert admin.post(
        f"/api/v1/publications/{pub['id']}/people",
        json={"person_id": reviewer_person["id"], "role": "reviewer"},
    ).status_code == 201
    assert [(m["To"], m["Subject"]) for m in sent] == [
        ("rae.reviewer@example.edu", "Review request: A Democratized Paper")
    ]
    # …and they now appear in the suggested acknowledgment.
    ack = editor.get(f"/api/v1/publications/{pub['id']}/acknowledgment").json()
    assert ack["reviewers"] == ["Rae Reviewer"]
    assert "Rae Reviewer" in ack["text"]

    # Office status changes notify the paper's editors.
    sent.clear()
    r = admin.post(f"/api/v1/publications/{pub['id']}/status", json={"status": "submitted"})
    assert r.status_code == 200, r.text
    assert [(m["To"], m["Subject"]) for m in sent] == [
        ("erin.editor@example.edu", "Publication status update: A Democratized Paper")
    ]
    assert "collab review to submitted" in sent[0].get_content()


def test_working_group_crud(admin):
    # Office creates a group.
    r = admin.post(
        "/api/v1/working-groups",
        json={"name": "Detector Simulation", "slug": "detector-sim", "description": "Sim work"},
    )
    assert r.status_code == 201, r.text
    wg = r.json()
    assert wg["slug"] == "detector-sim"
    assert wg["is_active"] is True
    assert wg["member_count"] == 0

    # Duplicate slug is rejected; malformed slug fails validation.
    assert admin.post(
        "/api/v1/working-groups", json={"name": "Dup", "slug": "detector-sim"}
    ).status_code == 409
    assert admin.post(
        "/api/v1/working-groups", json={"name": "Bad", "slug": "Not A Slug!"}
    ).status_code == 422

    # Office edits name/description/is_active.
    r = admin.patch(
        f"/api/v1/working-groups/{wg['id']}",
        json={"name": "Detector & Simulation", "is_active": False},
    )
    assert r.status_code == 200, r.text
    assert r.json()["name"] == "Detector & Simulation"
    assert r.json()["is_active"] is False
    assert admin.patch("/api/v1/working-groups/999999", json={"name": "x"}).status_code == 404

    # Members can see the list but cannot create or edit groups.
    member, pid = _linked_member(
        admin, given="Wanda", family="Groupless", email="wanda.groupless@example.edu"
    )
    assert member.get("/api/v1/working-groups").status_code == 200
    assert member.post(
        "/api/v1/working-groups", json={"name": "Rogue", "slug": "rogue"}
    ).status_code == 403
    assert member.patch(
        f"/api/v1/working-groups/{wg['id']}", json={"name": "Hijacked"}
    ).status_code == 403

    # A member may join a group themselves, but not enroll someone else.
    r = member.post(f"/api/v1/working-groups/{wg['id']}/members", json={"person_id": pid})
    assert r.status_code == 201, r.text
    # The membership shows up on the person detail (profile page display).
    detail = member.get(f"/api/v1/people/{pid}").json()
    assert [w["slug"] for w in detail["working_groups"]] == ["detector-sim"]
    assert member.post(
        f"/api/v1/working-groups/{wg['id']}/members", json={"person_id": pid}
    ).status_code == 409
    other = admin.post(
        "/api/v1/people/register",
        json={
            "given_name": "Otto",
            "family_name": "Other",
            "email": "otto.other@example.edu",
            "career_stage": "postdoc",
        },
    ).json()
    assert member.post(
        f"/api/v1/working-groups/{wg['id']}/members", json={"person_id": other["id"]}
    ).status_code == 403

    listed = admin.get("/api/v1/working-groups").json()
    assert {"slug": "detector-sim", "count": 1} in [
        {"slug": w["slug"], "count": w["member_count"]} for w in listed
    ]
    names = [p["family_name"] for p in admin.get(f"/api/v1/working-groups/{wg['id']}/members").json()]
    assert names == ["Groupless"]

    # A member may not remove someone else, but may leave themselves.
    assert admin.post(
        f"/api/v1/working-groups/{wg['id']}/members", json={"person_id": other["id"]}
    ).status_code == 201
    assert member.delete(
        f"/api/v1/working-groups/{wg['id']}/members/{other['id']}"
    ).status_code == 403
    assert member.delete(
        f"/api/v1/working-groups/{wg['id']}/members/{pid}"
    ).status_code == 204
    assert member.get(f"/api/v1/people/{pid}").json()["working_groups"] == []
    # Leaving again is a 404 — no longer a member.
    assert member.delete(
        f"/api/v1/working-groups/{wg['id']}/members/{pid}"
    ).status_code == 404

    # The office can remove anyone.
    assert admin.delete(
        f"/api/v1/working-groups/{wg['id']}/members/{other['id']}"
    ).status_code == 204
    listed = admin.get("/api/v1/working-groups").json()
    assert {"slug": "detector-sim", "count": 0} in [
        {"slug": w["slug"], "count": w["member_count"]} for w in listed
    ]


def test_member_stats(admin):
    # An active member with a research area and a US affiliation, so every
    # breakdown has at least one known row (the test stays self-sufficient).
    member, pid = _linked_member(
        admin, given="Ana", family="Areas", email="ana.areas@example.edu"
    )
    r = member.patch(
        f"/api/v1/people/{pid}",
        json={"research_areas": "Accelerator Physics", "usmcc_percent": 50},
    )
    assert r.status_code == 200, r.text
    inst = admin.post("/api/v1/institutions", json={"name": "Stats University"}).json()
    assert admin.post(
        f"/api/v1/people/{pid}/affiliations",
        json={"institution_id": inst["id"], "is_primary": True, "start_date": "2026-01-01"},
    ).status_code == 201

    r = admin.get("/api/v1/stats/members")
    assert r.status_code == 200, r.text
    stats = r.json()

    # The status breakdown covers every person record.
    by_status = {row["label"]: row["count"] for row in stats["by_status"]}
    assert sum(by_status.values()) == stats["total_people"] > 0
    assert stats["active"] == by_status.get("active", 0) > 0

    # career_stage is NOT NULL, so the stage breakdown covers all actives.
    assert sum(row["count"] for row in stats["by_career_stage"]) == stats["active"]

    # Everyone active in this run went through a recorded pending→active
    # transition, so the growth series accounts for at least all of them.
    assert sum(row["count"] for row in stats["new_members_by_month"]) >= stats["active"]
    # The series is a contiguous, ordered run of YYYY-MM buckets.
    months = [row["month"] for row in stats["new_members_by_month"]]
    assert months == sorted(months)
    assert all(len(m) == 7 and m[4] == "-" for m in months)

    # Voting members and US-based members are subsets of the active members
    # (non-voting is presented as active - voting, so the bound covers both).
    assert 0 <= stats["voting"] <= stats["active"]
    assert 1 <= stats["us_active"] <= stats["active"]
    assert stats["institutions_with_active"] >= 1

    # Effort aggregates cover the active members who reported usmcc_percent —
    # at least the 50% reported above.
    assert 1 <= stats["usmcc_reporting"] <= stats["active"]
    assert 0 < stats["avg_usmcc_percent"] <= 100
    # The percent-time buckets partition exactly the members reporting.
    assert sum(row["count"] for row in stats["by_usmcc_percent"]) == stats["usmcc_reporting"]
    by_pct = {row["label"]: row["count"] for row in stats["by_usmcc_percent"]}
    assert by_pct.get("50-100%", 0) >= 1  # the 50% reported above
    assert stats["usmcc_fte"] >= 0.5
    # avg × reporting and summed FTE describe the same underlying values.
    expected_fte = stats["avg_usmcc_percent"] * stats["usmcc_reporting"] / 100
    assert abs(stats["usmcc_fte"] - expected_fte) < 1e-6

    areas = {row["label"]: row["count"] for row in stats["by_research_area"]}
    assert areas.get("Accelerator Physics", 0) >= 1

    # Requires authentication.
    fresh = TestClient(app)
    assert fresh.get("/api/v1/stats/members").status_code == 401


def test_imported_members_counted_in_growth_stats(admin, tmp_path):
    # The CSV/XLSX importers create people directly as active, bypassing the
    # status-change endpoint; they must still record the transition-to-active
    # event that the "new members per month" chart is built from (issue #41).
    from app.cli import import_members

    csv_file = tmp_path / "members.csv"
    csv_file.write_text(
        "given_name,family_name,email,orcid,institution_short_name,"
        "career_stage,start_date,is_author\n"
        "Imre,Importson,imre.importson@example.edu,,ImpU,faculty,2019-05-01,false\n"
    )
    import_members(csv_path=csv_file, dry_run=False)

    people = admin.get("/api/v1/people", params={"q": "imre.importson"}).json()
    assert len(people) == 1
    pid = people[0]["id"]
    events = admin.get(f"/api/v1/people/{pid}/events").json()
    assert [(e["to_status"], e["effective_date"]) for e in events] == [
        ("active", "2019-05-01")
    ]

    # The growth series buckets the import under its start_date month.
    stats = admin.get("/api/v1/stats/members").json()
    months = {row["month"]: row["count"] for row in stats["new_members_by_month"]}
    assert months.get("2019-05", 0) >= 1

    # Re-importing (upsert) must not duplicate the activation event.
    import_members(csv_path=csv_file, dry_run=False)
    assert len(admin.get(f"/api/v1/people/{pid}/events").json()) == 1


def test_import_members_institution_move(admin, tmp_path):
    # Re-importing a member whose primary affiliation moved to another
    # institution must close the old open primary with the API's move
    # semantics instead of inserting a second open primary, which trips
    # uq_one_open_primary_affiliation and rolls back the whole import
    # (issue #54).
    from app.cli import import_members

    header = (
        "given_name,family_name,email,orcid,institution_short_name,"
        "career_stage,start_date,is_author\n"
    )

    def run(inst: str, start: str) -> None:
        csv_file = tmp_path / "members.csv"
        csv_file.write_text(
            header
            + f"Mona,Moverson,mona.moverson@example.edu,,{inst},faculty,{start},false\n"
        )
        import_members(csv_path=csv_file, dry_run=False)

    run("MovA", "2024-01-10")
    run("MovB", "2025-03-01")

    pid = admin.get("/api/v1/people", params={"q": "mona.moverson"}).json()[0]["id"]
    person = admin.get(f"/api/v1/people/{pid}").json()
    affils = sorted(person["affiliations"], key=lambda x: x["start_date"])
    assert [a["institution"]["short_name"] for a in affils] == ["MovA", "MovB"]
    # The old row ends ON the move date (inclusive on both ends, issue #67).
    assert affils[0]["end_date"] == "2025-03-01"
    assert affils[1]["end_date"] is None

    # A same-day move is a correction: the superseded row is deleted.
    run("MovC", "2025-03-01")
    person = admin.get(f"/api/v1/people/{pid}").json()
    open_affils = [x for x in person["affiliations"] if x["end_date"] is None]
    assert [x["institution"]["short_name"] for x in open_affils] == ["MovC"]
    assert all(x["institution"]["short_name"] != "MovB" for x in person["affiliations"])

    # A start date before the current affiliation's start can't be applied as
    # a move; the row's affiliation change is skipped, not stacked or crashed.
    run("MovD", "2020-01-01")
    person = admin.get(f"/api/v1/people/{pid}").json()
    open_affils = [x for x in person["affiliations"] if x["end_date"] is None]
    assert [x["institution"]["short_name"] for x in open_affils] == ["MovC"]

    # Re-importing at the current institution stays idempotent.
    run("MovC", "2026-01-01")
    person = admin.get(f"/api/v1/people/{pid}").json()
    assert len([x for x in person["affiliations"] if x["end_date"] is None]) == 1


def test_import_members_xlsx_institution_move(admin, tmp_path):
    # Same as above through the XLSX importer path (issue #54).
    from datetime import datetime

    import openpyxl

    from app.cli import import_members_xlsx

    header = [
        "Timestamp",
        "According to this definition, are you registering to be a voting "
        "or non-voting member of USMCC?",
        "First Name",
        "Middle Name",
        "Last Name",
        "Primary Affiliation",
        "Any additional affiliations",
        "Email",
        "ORCID ID (if available)",
        "Position",
        "Area(s) of Expertise",
        "What percent of your research time do you expect to spend on muon "
        "colliders in the next few years?",
    ]

    def run(inst: str, ts: "datetime") -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Members"
        ws.append(header)
        ws.append(
            [ts, "Voting member", "Xavier", None, "Xlsxson", inst, None,
             "xavier.xlsxson@example.edu", None, "Faculty",
             "Accelerator Physics", "10-24%"]
        )
        path = tmp_path / "members.xlsx"
        wb.save(path)
        import_members_xlsx(
            xlsx_path=path, sheet="Members", authors_from_voting=False, dry_run=False
        )

    run("Xlsx Institute A", datetime(2025, 4, 10))
    run("Xlsx Institute B", datetime(2025, 9, 2))

    pid = admin.get("/api/v1/people", params={"q": "xavier.xlsxson"}).json()[0]["id"]
    person = admin.get(f"/api/v1/people/{pid}").json()
    affils = sorted(person["affiliations"], key=lambda x: x["start_date"])
    assert [a["institution"]["name"] for a in affils] == [
        "Xlsx Institute A",
        "Xlsx Institute B",
    ]
    assert affils[0]["end_date"] == "2025-09-02"
    assert affils[1]["end_date"] is None


def test_import_members_xlsx_percent_time(admin, tmp_path):
    # The registration form asks for percent of research time as a range
    # ('10-24%'); the importer stores the midpoint in usmcc_percent, and
    # non-answers ('Too unsure to estimate here') stay unset.
    from datetime import datetime

    import openpyxl

    from app.cli import import_members_xlsx

    header = [
        "Timestamp",
        "According to this definition, are you registering to be a voting "
        "or non-voting member of USMCC?",
        "First Name",
        "Middle Name",
        "Last Name",
        "Primary Affiliation",
        "Any additional affiliations",
        "Email",
        "ORCID ID (if available)",
        "Position",
        "Area(s) of Expertise",
        "What percent of your research time do you expect to spend on muon "
        "colliders in the next few years?",
    ]

    def build(rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Members"
        ws.append(header)
        for r in rows:
            ws.append(r)
        path = tmp_path / "members.xlsx"
        wb.save(path)
        return path

    common = [datetime(2025, 4, 10), "Voting member"]
    path = build(
        [
            common
            + ["Pia", None, "Percentson", "PctU", None,
               "pia.percentson@example.edu", None, "Faculty",
               "Accelerator Physics", "10-24%"],
            common
            + ["Ulla", None, "Unsure", "PctU", None,
               "ulla.unsure@example.edu", None, "Postdoc",
               "Accelerator Physics", "Too unsure to estimate here"],
        ]
    )
    import_members_xlsx(
        xlsx_path=path, sheet="Members", authors_from_voting=False, dry_run=False
    )

    pia = admin.get("/api/v1/people", params={"q": "pia.percentson"}).json()[0]
    assert admin.get(f"/api/v1/people/{pia['id']}").json()["usmcc_percent"] == 17
    ulla = admin.get("/api/v1/people", params={"q": "ulla.unsure"}).json()[0]
    assert admin.get(f"/api/v1/people/{ulla['id']}").json()["usmcc_percent"] is None

    # Re-import (upsert): a real answer updates the stored value, and a
    # non-answer must not wipe one that exists.
    path = build(
        [
            common
            + ["Pia", None, "Percentson", "PctU", None,
               "pia.percentson@example.edu", None, "Faculty",
               "Accelerator Physics", "Too unsure to estimate here"],
            common
            + ["Ulla", None, "Unsure", "PctU", None,
               "ulla.unsure@example.edu", None, "Postdoc",
               "Accelerator Physics", "50-100%"],
        ]
    )
    import_members_xlsx(
        xlsx_path=path, sheet="Members", authors_from_voting=False, dry_run=False
    )
    assert admin.get(f"/api/v1/people/{pia['id']}").json()["usmcc_percent"] == 17
    assert admin.get(f"/api/v1/people/{ulla['id']}").json()["usmcc_percent"] == 75


def test_import_members_xlsx_voting_eligibility(admin, tmp_path):
    # A blank voting answer must not grant the flag, and explicit voting
    # requests are subject to the same charter rules the API enforces:
    # active, non-student, currently at a US institution (issue #58).
    from datetime import datetime

    import openpyxl

    from app.cli import import_members_xlsx

    header = [
        "Timestamp",
        "According to this definition, are you registering to be a voting "
        "or non-voting member of USMCC?",
        "First Name",
        "Middle Name",
        "Last Name",
        "Primary Affiliation",
        "Any additional affiliations",
        "Email",
        "ORCID ID (if available)",
        "Position",
        "Area(s) of Expertise",
        "What percent of your research time do you expect to spend on muon "
        "colliders in the next few years?",
    ]

    admin.post(
        "/api/v1/institutions",
        json={"name": "Xlsx Overseas Institute", "country": "Japan", "is_us": False},
    )

    def run(rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Members"
        ws.append(header)
        for r in rows:
            ws.append(r)
        path = tmp_path / "members.xlsx"
        wb.save(path)
        import_members_xlsx(
            xlsx_path=path, sheet="Members", authors_from_voting=True, dry_run=False
        )

    ts = datetime(2025, 4, 10)
    run(
        [
            # Blank answer: imports as non-voting, not voting.
            [ts, None, "Blanche", None, "Blankvote", "Xlsx Voting University", None,
             "blanche.blankvote@example.edu", None, "Faculty",
             "Accelerator Physics", "10-24%"],
            # A grad student's voting request is dropped (charter rule).
            [ts, "Voting member", "Stu", None, "Studentvote", "Xlsx Voting University",
             None, "stu.studentvote@example.edu", None, "Graduate Student",
             "Detector R&D", "10-24%"],
            # A voting request from a non-US institution is dropped too.
            [ts, "Voting member", "Oli", None, "Overseasvote", "Xlsx Overseas Institute",
             None, "oli.overseasvote@example.edu", None, "Faculty",
             "Accelerator Physics", "10-24%"],
            # An eligible request is granted (and opens an author period).
            [ts, "Voting member", "Val", None, "Validvote", "Xlsx Voting University",
             None, "val.validvote@example.edu", None, "Faculty",
             "Accelerator Physics", "10-24%"],
        ]
    )

    def fetch(q):
        pid = admin.get("/api/v1/people", params={"q": q}).json()[0]["id"]
        return admin.get(f"/api/v1/people/{pid}").json()

    assert fetch("blanche.blankvote")["is_voting"] is False
    assert fetch("stu.studentvote")["is_voting"] is False
    assert fetch("oli.overseasvote")["is_voting"] is False
    val = fetch("val.validvote")
    assert val["is_voting"] is True

    # authors_from_voting only opens periods for members actually granted
    # the flag.
    assert admin.get(f"/api/v1/people/{val['id']}/author-periods").json() != []
    ineligible = fetch("stu.studentvote")
    assert admin.get(f"/api/v1/people/{ineligible['id']}/author-periods").json() == []

    # Re-import (upsert) with a blank answer keeps the stored flag rather
    # than silently revoking (or granting) voting membership.
    run(
        [
            [ts, None, "Val", None, "Validvote", "Xlsx Voting University", None,
             "val.validvote@example.edu", None, "Faculty",
             "Accelerator Physics", "10-24%"],
        ]
    )
    assert fetch("val.validvote")["is_voting"] is True
    # …and an explicit "non-voting" answer clears it.
    run(
        [
            [ts, "Non-voting member", "Val", None, "Validvote", "Xlsx Voting University",
             None, "val.validvote@example.edu", None, "Faculty",
             "Accelerator Physics", "10-24%"],
        ]
    )
    assert fetch("val.validvote")["is_voting"] is False


def test_import_members_warns_on_bad_start_date(admin, tmp_path, capsys):
    # An unparseable start_date must not silently become today with no trace
    # (issue #66): the row still imports, but with a per-row warning.
    from app.cli import import_members

    csv_file = tmp_path / "members.csv"
    csv_file.write_text(
        "given_name,family_name,email,orcid,institution_short_name,"
        "career_stage,start_date,is_author\n"
        "Dana,Dateson,dana.dateson@example.edu,,DateU,faculty,07/15/2024,false\n"
    )
    import_members(csv_path=csv_file, dry_run=False)
    out = capsys.readouterr().out
    assert "unparseable start_date '07/15/2024'" in out
    assert len(admin.get("/api/v1/people", params={"q": "dana.dateson"}).json()) == 1


def test_import_members_xlsx_warns_on_bad_timestamp(admin, tmp_path, capsys):
    # A non-date Timestamp cell must not silently become today either
    # (issue #66).
    import openpyxl

    from app.cli import import_members_xlsx

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Members"
    ws.append(
        ["Timestamp",
         "According to this definition, are you registering to be a voting "
         "or non-voting member of USMCC?",
         "First Name", "Middle Name", "Last Name", "Primary Affiliation",
         "Any additional affiliations", "Email", "ORCID ID (if available)",
         "Position", "Area(s) of Expertise",
         "What percent of your research time do you expect to spend on muon "
         "colliders in the next few years?"]
    )
    ws.append(
        ["07/15/2024", "Non-voting member", "Tim", None, "Stampson",
         "Stamp University", None, "tim.stampson@example.edu", None,
         "Faculty", "Accelerator Physics", "10-24%"]
    )
    path = tmp_path / "members.xlsx"
    wb.save(path)
    import_members_xlsx(
        xlsx_path=path, sheet="Members", authors_from_voting=False, dry_run=False
    )
    out = capsys.readouterr().out
    assert "Timestamp '07/15/2024' is not a date cell" in out
    assert len(admin.get("/api/v1/people", params={"q": "tim.stampson"}).json()) == 1


def test_percent_time_fraction_cells():
    # Excel stores a cell *displayed* as "25%" as the float 0.25, which used
    # to round to 0 and record 0% effort (issue #66); fractions in (0, 1]
    # scale back to percent. Range strings and plain numbers are unchanged.
    from app.cli import _percent_time

    assert _percent_time(0.25) == 25
    assert _percent_time(0.5) == 50
    assert _percent_time(1) == 100  # a percent-formatted 100%
    assert _percent_time(0) == 0
    assert _percent_time(50) == 50
    assert _percent_time("0-10%") == 5
    assert _percent_time(150) is None
    assert _percent_time(None) is None


def test_import_members_infile_duplicate_rows(admin, tmp_path):
    # The session is autoflush=False, so pending rows used to be invisible to
    # the dedup SELECTs: the same email twice in one CSV produced a second
    # activation event (double-counting growth stats), a second open primary
    # affiliation, and a second author period — the latter two roll back the
    # whole import on their constraints (issue #66).
    from app.cli import import_members

    csv_file = tmp_path / "members.csv"
    row = "Dupla,Duplison,dupla.duplison@example.edu,,DupU,faculty,2024-02-01,true\n"
    csv_file.write_text(
        "given_name,family_name,email,orcid,institution_short_name,"
        "career_stage,start_date,is_author\n" + row + row
    )
    import_members(csv_path=csv_file, dry_run=False)

    pid = admin.get("/api/v1/people", params={"q": "dupla.duplison"}).json()[0]["id"]
    events = admin.get(f"/api/v1/people/{pid}/events").json()
    assert [e["to_status"] for e in events] == ["active"]
    assert len(admin.get(f"/api/v1/people/{pid}/author-periods").json()) == 1
    person = admin.get(f"/api/v1/people/{pid}").json()
    assert len([a for a in person["affiliations"] if a["end_date"] is None]) == 1


def test_import_talks_xlsx_infile_duplicate_rows(admin, tmp_path):
    # Same autoflush=False pitfall in the talks importer: the duplicate
    # SELECT must see talks pending from earlier rows of the same file
    # (issue #66).
    from datetime import datetime

    import openpyxl

    from app.cli import import_talks_xlsx

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assigned Talks"
    ws.append(["Date", "Conference", "Topic", "Name", "Plenary/Parallel",
               "Invited/Contributed", "URL", "Notes"])
    row = [datetime(2025, 5, 6), "DupConf 2025", "Duplicated Topic", None,
           "Plenary", "Invited", None, None]
    ws.append(row)
    ws.append(row)
    path = tmp_path / "talks.xlsx"
    wb.save(path)
    import_talks_xlsx(xlsx_path=path, sheet="Assigned Talks", dry_run=False)

    talks = admin.get("/api/v1/talks").json()
    assert len([t for t in talks if t["title"] == "Duplicated Topic"]) == 1


def test_publication_short_code_survives_gaps(admin):
    # _next_short_code must allocate one past the highest existing suffix; a
    # row count re-issues an existing code once the sequence has gaps and the
    # unique constraint turns the next create into a 500 (issue #66).
    from sqlalchemy import text

    from app.db import engine

    first = admin.post(
        "/api/v1/publications", json={"title": "Gap One", "pub_type": "proceedings"}
    ).json()
    second = admin.post(
        "/api/v1/publications", json={"title": "Gap Two", "pub_type": "proceedings"}
    ).json()
    assert second["short_code"] > first["short_code"]

    with engine.begin() as conn:
        conn.execute(text("DELETE FROM publications WHERE id = :id"), {"id": first["id"]})

    r = admin.post(
        "/api/v1/publications", json={"title": "Gap Three", "pub_type": "proceedings"}
    )
    assert r.status_code == 201, r.text
    assert r.json()["short_code"] > second["short_code"]


def test_publication_short_code_collision_retries(admin, monkeypatch):
    # Two concurrent creates can compute the same code; the loser must retry
    # with a fresh code instead of surfacing the IntegrityError as a 500
    # (issue #66).
    import app.routers.publications as pubs_router

    first = admin.post(
        "/api/v1/publications", json={"title": "Race One", "pub_type": "note"}
    ).json()

    real = pubs_router._next_short_code
    calls = []

    def collide_once(db, pub_type):
        calls.append(pub_type)
        if len(calls) == 1:
            return first["short_code"]
        return real(db, pub_type)

    monkeypatch.setattr(pubs_router, "_next_short_code", collide_once)
    r = admin.post("/api/v1/publications", json={"title": "Race Two", "pub_type": "note"})
    assert r.status_code == 201, r.text
    assert r.json()["short_code"] != first["short_code"]
    assert len(calls) == 2


# --- ORCID sign-in & registration approval (issue #50) -------------------------


def _orcid_signin(monkeypatch, orcid_id, name):
    """Drive the full OAuth round trip with a faked code exchange; returns a
    fresh client (carrying whatever cookie the callback set) and the
    redirect. Starting at /orcid/login matters: the callback only accepts a
    state whose nonce matches the cookie minted there (issue #62)."""
    from urllib.parse import parse_qs, urlparse

    from app.routers import auth as auth_router

    async def fake_exchange(code, redirect_uri):
        return {"orcid": orcid_id, "name": name}

    monkeypatch.setattr(auth_router.orcid_svc, "exchange_code", fake_exchange)
    c = TestClient(app)
    r = c.get("/api/v1/auth/orcid/login", follow_redirects=False)
    assert r.status_code == 307, r.text
    state = parse_qs(urlparse(r.headers["location"]).query)["state"][0]
    r = c.get(
        f"/api/v1/auth/orcid/callback?code=fake&state={state}", follow_redirects=False
    )
    return c, r


def _no_session_cookie(r) -> bool:
    """True when the response granted no login session. (The callback always
    sends one Set-Cookie clearing its single-use state cookie, so 'no
    set-cookie at all' is not the right check.)"""
    return not any(
        c.startswith("usmccdb_session=") for c in r.headers.get_list("set-cookie")
    )


def test_orcid_stranger_gets_no_access_until_approved(admin, monkeypatch):
    import app.services.email as email_mod

    sent = []
    monkeypatch.setattr(email_mod, "_deliver", sent.append)

    stranger, r = _orcid_signin(monkeypatch, "0000-0002-1825-0097", "Josiah Carberry")
    assert r.status_code == 307, r.text
    assert r.headers["location"] == "/register?welcome=orcid"

    # Signed in, but the pending stub grants no member-level access (issue
    # #50: a free ORCID iD must not open the member directory or statistics).
    assert stranger.get("/api/v1/auth/me").status_code == 403
    assert stranger.get("/api/v1/people").status_code == 403
    assert stranger.get("/api/v1/stats/members").status_code == 403
    assert stranger.get("/api/v1/working-groups").status_code == 403

    # The sign-in left a pending person stub with an audit event.
    pend = admin.get(
        "/api/v1/people", params={"status": "pending", "q": "0000-0002-1825-0097"}
    ).json()
    assert len(pend) == 1
    pid = pend[0]["id"]
    assert pend[0]["orcid"] == "0000-0002-1825-0097"

    # Completing the registration form updates that same record — no
    # duplicate person, ORCID iD kept from the authenticated sign-in.
    r = stranger.post(
        "/api/v1/people/register",
        json={
            "given_name": "Josiah",
            "family_name": "Carberry",
            "email": "josiah.carberry@example.edu",
            "career_stage": "faculty",
            "institution_name": "Brown University",
            "institution_is_us": True,
        },
    )
    assert r.status_code == 201, r.text
    # Non-office registration answers with the neutral acknowledgement only
    # (issue #62); the record itself shows the completed registration.
    assert set(r.json()) == {"detail"}
    completed = admin.get(f"/api/v1/people/{pid}").json()
    assert completed["email"] == "josiah.carberry@example.edu"
    assert completed["orcid"] == "0000-0002-1825-0097"
    assert completed["status"] == "pending"
    events = admin.get(f"/api/v1/people/{pid}/events").json()
    assert [e["to_status"] for e in events] == ["pending"]

    # The submission asked the approvers to review it.
    assert len(sent) == 1
    assert "office@example.edu" in sent[0]["To"]
    assert sent[0]["Subject"] == "New membership registration: Josiah Carberry"

    # Still no access while pending; a fresh ORCID sign-in parks at the login
    # page without a session.
    assert stranger.get("/api/v1/people").status_code == 403
    parked, r = _orcid_signin(monkeypatch, "0000-0002-1825-0097", "Josiah Carberry")
    assert r.headers["location"] == "/login?error=membership_pending"
    assert _no_session_cookie(r)

    # Approval opens the door: the existing session works immediately and a
    # new ORCID sign-in lands on the home page.
    assert admin.post(
        f"/api/v1/people/{pid}/status", json={"status": "active"}
    ).status_code == 200
    assert stranger.get("/api/v1/people").status_code == 200
    assert stranger.get("/api/v1/auth/me").json()["person_id"] == pid
    approved, r = _orcid_signin(monkeypatch, "0000-0002-1825-0097", "Josiah Carberry")
    assert r.headers["location"] == "/"
    assert approved.get("/api/v1/people").status_code == 200


def test_orcid_rejected_registration_turned_away(admin, monkeypatch):
    import app.services.email as email_mod

    monkeypatch.setattr(email_mod, "_deliver", lambda msg: None)

    stranger, r = _orcid_signin(monkeypatch, "0000-0003-1111-2222", "Rae Jected")
    assert r.headers["location"] == "/register?welcome=orcid"
    r = stranger.post(
        "/api/v1/people/register",
        json={"given_name": "Rae", "family_name": "Jected", "email": "rae.jected@example.edu"},
    )
    assert r.status_code == 201, r.text
    pid = admin.get("/api/v1/people", params={"q": "rae.jected@example.edu"}).json()[0]["id"]
    admin.post(f"/api/v1/people/{pid}/status", json={"status": "rejected"})

    # The rejection cuts off the existing session and future sign-ins alike.
    assert stranger.get("/api/v1/people").status_code == 403
    _, r = _orcid_signin(monkeypatch, "0000-0003-1111-2222", "Rae Jected")
    assert r.headers["location"] == "/login?error=membership_rejected"
    assert _no_session_cookie(r)


def test_orcid_links_existing_approved_member(admin, monkeypatch):
    import app.services.email as email_mod

    monkeypatch.setattr(email_mod, "_deliver", lambda msg: None)

    person = admin.post(
        "/api/v1/people/register",
        json={
            "given_name": "Ora",
            "family_name": "Linked",
            "email": "ora.linked@example.edu",
            "orcid": "0000-0001-2345-6789",
        },
    ).json()
    admin.post(f"/api/v1/people/{person['id']}/status", json={"status": "active"})

    # First ORCID sign-in auto-links a login to the approved person record.
    member, r = _orcid_signin(monkeypatch, "0000-0001-2345-6789", "Ora Linked")
    assert r.headers["location"] == "/"
    me = member.get("/api/v1/auth/me")
    assert me.status_code == 200, me.text
    assert me.json()["person_id"] == person["id"]
    assert member.get("/api/v1/people").status_code == 200


def test_admin_contact_approves_pending_registration(admin, monkeypatch):
    import app.services.email as email_mod

    sent = []
    monkeypatch.setattr(email_mod, "_deliver", sent.append)

    inst = admin.post("/api/v1/institutions", json={"name": "Approve University"}).json()
    other_inst = admin.post("/api/v1/institutions", json={"name": "Far Away Tech"}).json()
    contact, contact_pid = _linked_member(
        admin, given="Cal", family="Contact", email="cal.contact@example.edu",
        career_stage="staff",
    )
    assert admin.post(
        f"/api/v1/people/{contact_pid}/affiliations",
        json={"institution_id": inst["id"], "is_primary": True, "start_date": "2025-01-01"},
    ).status_code == 201
    assert admin.post(
        "/api/v1/collab-roles",
        json={
            "person_id": contact_pid,
            "role": "admin_contact",
            "institution_id": inst["id"],
            "start_date": "2025-01-01",
        },
    ).status_code == 201

    # A stranger registers at the contact's institution.
    sent.clear()  # drop mail generated by the setup above
    registrant = TestClient(app)
    r = registrant.post(
        "/api/v1/people/register",
        json={
            "given_name": "New",
            "family_name": "Comer",
            "email": "new.comer@example.edu",
            "career_stage": "grad",
            "institution_id": inst["id"],
        },
    )
    assert r.status_code == 201, r.text
    pid = admin.get("/api/v1/people", params={"q": "new.comer@example.edu"}).json()[0]["id"]

    # The approval request went to the office and the institution's contact.
    assert len(sent) == 1
    assert "office@example.edu" in sent[0]["To"]
    assert "cal.contact@example.edu" in sent[0]["To"]

    # The admin contact may approve the pending registration…
    assert contact.post(
        f"/api/v1/people/{pid}/status", json={"status": "active"}
    ).status_code == 200
    # …but gets no say over regular status changes afterwards…
    assert contact.post(
        f"/api/v1/people/{pid}/status", json={"status": "inactive"}
    ).status_code == 403
    # …or over pending registrations at other institutions.
    outsider = TestClient(app)
    r = outsider.post(
        "/api/v1/people/register",
        json={
            "given_name": "Els",
            "family_name": "Ewhere",
            "email": "els.ewhere@example.edu",
            "institution_id": other_inst["id"],
        },
    )
    assert r.status_code == 201, r.text
    els_id = admin.get("/api/v1/people", params={"q": "els.ewhere@example.edu"}).json()[0]["id"]
    assert contact.post(
        f"/api/v1/people/{els_id}/status", json={"status": "rejected"}
    ).status_code == 403


# --- Auth hardening (issue #62) -------------------------------------------------


def test_login_gives_one_generic_failure_answer(admin):
    """Unknown username, wrong password, and disabled account must be
    indistinguishable — a distinct answer confirms the account exists."""
    r = admin.post(
        "/api/v1/auth/users",
        json={"username": "dis.abled", "password": "pw-123456", "role": "member"},
    )
    assert r.status_code == 201, r.text
    uid = r.json()["id"]
    assert admin.patch(f"/api/v1/auth/users/{uid}", json={"is_active": False}).status_code == 200

    c = TestClient(app)
    unknown = c.post(
        "/api/v1/auth/login", json={"username": "no.such.user", "password": "whatever"}
    )
    wrong = c.post(
        "/api/v1/auth/login", json={"username": "dis.abled", "password": "not-the-pw"}
    )
    disabled = c.post(
        "/api/v1/auth/login", json={"username": "dis.abled", "password": "pw-123456"}
    )
    assert unknown.status_code == wrong.status_code == disabled.status_code == 401
    assert unknown.json() == wrong.json() == disabled.json()


def test_orcid_callback_rejects_state_from_another_session(admin, monkeypatch):
    """A signed state minted in one browser must not sign in another browser:
    an attacker could otherwise hand their own callback URL to a victim and
    silently log the victim into the attacker's account (login CSRF)."""
    from urllib.parse import parse_qs, urlparse

    from app.routers import auth as auth_router

    async def fake_exchange(code, redirect_uri):
        return {"orcid": "0000-0002-0000-0001", "name": "Eve Attacker"}

    monkeypatch.setattr(auth_router.orcid_svc, "exchange_code", fake_exchange)

    # The attacker starts a sign-in and captures their callback URL…
    attacker = TestClient(app)
    r = attacker.get("/api/v1/auth/orcid/login", follow_redirects=False)
    state = parse_qs(urlparse(r.headers["location"]).query)["state"][0]

    # …but the victim's browser carries no matching state cookie.
    victim = TestClient(app)
    r = victim.get(
        f"/api/v1/auth/orcid/callback?code=fake&state={state}", follow_redirects=False
    )
    assert r.headers["location"] == "/login?error=orcid_state"
    assert _no_session_cookie(r)

    # A validly-signed state without a nonce (the pre-fix format) fails too.
    legacy_state = auth_router._state_serializer().dumps({"next": "/"})
    r = victim.get(
        f"/api/v1/auth/orcid/callback?code=fake&state={legacy_state}",
        follow_redirects=False,
    )
    assert r.headers["location"] == "/login?error=orcid_state"
    assert _no_session_cookie(r)


def test_registration_does_not_reveal_existing_records(admin, monkeypatch):
    """Anonymous registration answers identically whether or not the email /
    ORCID iD already belongs to a member; the office is told about the
    duplicate instead. Office callers keep the informative 409."""
    import app.services.email as email_mod

    sent = []
    monkeypatch.setattr(email_mod, "_deliver", sent.append)

    existing = admin.post(
        "/api/v1/people/register",
        json={
            "given_name": "Dana",
            "family_name": "Duplicated",
            "email": "dana.duplicated@example.edu",
            "orcid": "0000-0002-4444-5555",
        },
    )
    assert existing.status_code == 201, existing.text

    anon = TestClient(app)
    sent.clear()

    # Same email as the existing member…
    dup = anon.post(
        "/api/v1/people/register",
        json={
            "given_name": "Pat",
            "family_name": "Probe",
            "email": "dana.duplicated@example.edu",
        },
    )
    # …same ORCID iD…
    dup_orcid = anon.post(
        "/api/v1/people/register",
        json={
            "given_name": "Pat",
            "family_name": "Probe",
            "email": "pat.probe.orcid@example.edu",
            "orcid": "0000-0002-4444-5555",
        },
    )
    # …and a genuinely new person all get the exact same answer.
    fresh = anon.post(
        "/api/v1/people/register",
        json={
            "given_name": "Frida",
            "family_name": "Fresh",
            "email": "frida.fresh@example.edu",
        },
    )
    assert dup.status_code == dup_orcid.status_code == fresh.status_code == 201
    assert dup.json() == dup_orcid.json() == fresh.json()
    assert set(fresh.json()) == {"detail"}

    # The probes created nothing; the real registration went through.
    assert admin.get("/api/v1/people", params={"q": "Probe"}).json() == []
    assert len(admin.get("/api/v1/people", params={"q": "dana.duplicated"}).json()) == 1
    assert len(admin.get("/api/v1/people", params={"q": "frida.fresh"}).json()) == 1

    # The office heard about both duplicate attempts (and the fresh one).
    subjects = [m["Subject"] for m in sent]
    assert subjects.count("Duplicate membership registration: Pat Probe") == 2
    assert "New membership registration: Frida Fresh" in subjects
    dup_mail = next(m for m in sent if m["Subject"].startswith("Duplicate"))
    assert "office@example.edu" in dup_mail["To"]
    assert "Dana Duplicated" in dup_mail.get_content()

    # Office/admin callers still get the informative 409.
    r = admin.post(
        "/api/v1/people/register",
        json={
            "given_name": "Dana",
            "family_name": "Duplicated",
            "email": "dana.duplicated@example.edu",
        },
    )
    assert r.status_code == 409
    assert "already exists" in r.json()["detail"]


def test_login_rate_limited_per_ip(admin, monkeypatch):
    from app import ratelimit

    limiter = ratelimit.login_limiter()
    monkeypatch.setattr(limiter, "limit", 3)
    limiter.reset()
    try:
        c = TestClient(app)
        bad = {"username": "no.such.user", "password": "guess"}
        for _ in range(3):
            assert c.post("/api/v1/auth/login", json=bad).status_code == 401
        r = c.post("/api/v1/auth/login", json=bad)
        assert r.status_code == 429
        assert "retry-after" in r.headers
        # Existing sessions are untouched — only the login endpoint is gated.
        assert admin.get("/api/v1/auth/me").status_code == 200
    finally:
        limiter.reset()


def test_registration_rate_limited_for_non_office(admin, monkeypatch):
    import app.services.email as email_mod
    from app import ratelimit

    monkeypatch.setattr(email_mod, "_deliver", lambda msg: None)
    limiter = ratelimit.registration_limiter()
    monkeypatch.setattr(limiter, "limit", 2)
    limiter.reset()
    try:
        anon = TestClient(app)

        def payload(i):
            return {
                "given_name": "Rate",
                "family_name": f"Limited{i}",
                "email": f"rate.limited{i}@example.edu",
            }

        assert anon.post("/api/v1/people/register", json=payload(1)).status_code == 201
        assert anon.post("/api/v1/people/register", json=payload(2)).status_code == 201
        r = anon.post("/api/v1/people/register", json=payload(3))
        assert r.status_code == 429
        assert "retry-after" in r.headers
        # Signed-in office accounts are not rate limited.
        assert admin.post("/api/v1/people/register", json=payload(4)).status_code == 201
    finally:
        limiter.reset()
