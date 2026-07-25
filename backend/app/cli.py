"""Management CLI.

    python -m app.cli create-admin --username chair
    python -m app.cli import-members members.csv [--dry-run]
    python -m app.cli seed-wgs

CSV columns (header required):
    given_name, family_name, email, orcid, institution_short_name,
    career_stage, start_date (YYYY-MM-DD), is_author (true/false)
"""

import csv
import re
from datetime import date
from pathlib import Path

import typer
from sqlalchemy import select

from app.db import SessionLocal
from app.models.membership import RESEARCH_AREAS
from app.models import (
    Affiliation,
    AuthorPeriod,
    CareerStage,
    CollabRole,
    CollabRoleType,
    Institution,
    MembershipEvent,
    MemberStatus,
    Person,
    User,
    UserRole,
    WorkingGroup,
)
from app.security import hash_password

cli = typer.Typer(help="USMCC database management commands")

DEFAULT_WGS = [
    ("Accelerator", "accelerator"),
    ("Detector", "detector"),
    ("Physics", "physics"),
    ("Community Engagement", "community"),
]


@cli.command()
def create_admin(
    username: str = typer.Option(..., prompt=True),
    password: str = typer.Option(..., prompt=True, hide_input=True, confirmation_prompt=True),
):
    """Create (or reset the password of) a local admin account."""
    with SessionLocal() as db:
        user = db.execute(select(User).where(User.username == username)).scalar_one_or_none()
        if user is None:
            user = User(username=username, role=UserRole.admin)
            db.add(user)
            typer.echo(f"Creating admin '{username}'")
        else:
            user.role = UserRole.admin
            typer.echo(f"Updating existing user '{username}' (now admin, password reset)")
        user.password_hash = hash_password(password)
        user.is_active = True
        db.commit()
    typer.echo("Done.")


@cli.command()
def seed_wgs():
    """Insert the initial USMCC working groups (idempotent)."""
    with SessionLocal() as db:
        for name, slug in DEFAULT_WGS:
            if db.execute(
                select(WorkingGroup).where(WorkingGroup.slug == slug)
            ).scalar_one_or_none():
                continue
            db.add(WorkingGroup(name=name, slug=slug))
            typer.echo(f"Added working group: {name}")
        db.commit()
    typer.echo("Done.")


def _ensure_activation_event(db, person: Person, effective: date) -> None:
    """Record the transition to active for an imported member, if no such
    event exists yet. Imports create people directly as active, bypassing the
    status-change endpoint that normally writes the audit trail — without this
    the "new members per month" statistics never see imported members."""
    if person.status != MemberStatus.active:
        return
    has_event = db.execute(
        select(MembershipEvent.id).where(
            MembershipEvent.person_id == person.id,
            MembershipEvent.to_status == MemberStatus.active.value,
        )
    ).first()
    if has_event is None:
        db.add(
            MembershipEvent(
                person_id=person.id,
                from_status=None,
                to_status=MemberStatus.active.value,
                effective_date=effective,
                note="Imported member record",
            )
        )


def _set_primary_affiliation(
    db, ref: str, person: Person, inst: Institution, stage: CareerStage, start: date
) -> None:
    """Ensure the person's open primary affiliation is at ``inst``, applying
    the same move semantics as the API (``_close_primary`` in
    app/routers/people.py): an open primary at another institution is closed
    ON ``start`` (date ranges are inclusive on both ends, so the person
    carries both affiliations on the transition day), and a same-day move
    deletes the superseded row instead."""
    open_primary = db.execute(
        select(Affiliation).where(
            Affiliation.person_id == person.id,
            Affiliation.is_primary.is_(True),
            Affiliation.end_date.is_(None),
        )
    ).scalar_one_or_none()
    if open_primary is not None:
        if open_primary.institution_id == inst.id:
            return
        if start < open_primary.start_date:
            typer.echo(
                f"{ref}: SKIP institution move (start {start.isoformat()} predates "
                f"current affiliation start {open_primary.start_date.isoformat()})"
            )
            return
        if start == open_primary.start_date:
            db.delete(open_primary)
            # Flush now: the unit of work runs INSERTs before DELETEs, so the
            # replacement row would otherwise trip uq_one_open_primary_affiliation
            # while this one still exists.
            db.flush()
        else:
            open_primary.end_date = start
    elif db.execute(
        select(Affiliation.id).where(
            Affiliation.person_id == person.id,
            Affiliation.institution_id == inst.id,
            Affiliation.end_date.is_(None),
        )
    ).first():
        # No open primary, but an open (non-primary) row already ties the
        # person to this institution — don't stack a second open row.
        return
    db.add(
        Affiliation(
            person_id=person.id,
            institution_id=inst.id,
            is_primary=True,
            career_stage=stage,
            start_date=start,
        )
    )


@cli.command()
def import_members(
    csv_path: Path = typer.Argument(..., exists=True, readable=True),
    dry_run: bool = typer.Option(False, help="Parse and report, write nothing"),
):
    """Import an initial member list from CSV (upsert on email)."""
    created = updated = skipped = 0
    with SessionLocal() as db, open(csv_path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        for i, row in enumerate(reader, start=2):  # header is line 1
            email = (row.get("email") or "").strip().lower()
            given = (row.get("given_name") or "").strip()
            family = (row.get("family_name") or "").strip()
            if not (email and given and family):
                typer.echo(f"line {i}: SKIP (missing given_name/family_name/email)")
                skipped += 1
                continue

            orcid = (row.get("orcid") or "").strip() or None
            stage_raw = (row.get("career_stage") or "other").strip().lower()
            try:
                stage = CareerStage(stage_raw)
            except ValueError:
                typer.echo(f"line {i}: unknown career_stage '{stage_raw}', using 'other'")
                stage = CareerStage.other
            try:
                start = date.fromisoformat((row.get("start_date") or "").strip())
            except ValueError:
                start = date.today()
            is_author = (row.get("is_author") or "").strip().lower() in ("true", "yes", "1")

            person = db.execute(select(Person).where(Person.email == email)).scalar_one_or_none()
            if person is None:
                person = Person(
                    given_name=given,
                    family_name=family,
                    email=email,
                    orcid=orcid,
                    career_stage=stage,
                    status=MemberStatus.active,
                )
                db.add(person)
                db.flush()
                created += 1
            else:
                person.given_name, person.family_name = given, family
                person.orcid = orcid or person.orcid
                person.career_stage = stage
                updated += 1
            _ensure_activation_event(db, person, start)

            short = (row.get("institution_short_name") or "").strip()
            if short:
                inst = db.execute(
                    select(Institution).where(Institution.short_name == short)
                ).scalar_one_or_none()
                if inst is None:
                    inst = Institution(name=short, short_name=short)
                    db.add(inst)
                    db.flush()
                    typer.echo(f"line {i}: created institution '{short}' (fill in details later)")
                _set_primary_affiliation(db, f"line {i}", person, inst, stage, start)

            if is_author:
                has_period = db.execute(
                    select(AuthorPeriod).where(
                        AuthorPeriod.person_id == person.id,
                        AuthorPeriod.end_date.is_(None),
                    )
                ).scalar_one_or_none()
                if has_period is None:
                    db.add(AuthorPeriod(person_id=person.id, start_date=start))

        if dry_run:
            db.rollback()
            typer.echo(f"DRY RUN — would create {created}, update {updated}, skip {skipped}")
        else:
            db.commit()
            typer.echo(f"Imported: {created} created, {updated} updated, {skipped} skipped")


# --- Demo data -------------------------------------------------------------------

DEMO_INSTITUTIONS = [
    ("University of Springfield", "USpring", "University of Springfield, Springfield, IL 62901, USA"),
    ("Lakeview National Laboratory", "LNL", "Lakeview National Laboratory, Lakeview, IL 60510, USA"),
    ("Coastal State University", "CSU", "Coastal State University, Santa Rosa, CA 95401, USA"),
    ("Midwest Institute of Technology", "MIT-W", "Midwest Institute of Technology, Des Moines, IA 50309, USA"),
    ("Bayside University", "Bayside", "Bayside University, Gulfport, MS 39501, USA"),
    ("Northern Plains University", "NPU", "Northern Plains University, Fargo, ND 58102, USA"),
]

# (given, family, stage, voting, institution index, orcid)
DEMO_PEOPLE = [
    ("Maria", "Alvarez", CareerStage.faculty, True, 0, "0000-0002-0000-0001"),
    ("James", "O'Connor", CareerStage.faculty, True, 1, "0000-0002-0000-0002"),
    ("Wei", "Chen", CareerStage.postdoc, True, 1, "0000-0002-0000-0003"),
    ("Priya", "Sharma", CareerStage.faculty, True, 2, None),
    ("Sam", "Taylor", CareerStage.grad, False, 0, None),
    ("Fatima", "Hassan", CareerStage.staff, True, 1, "0000-0002-0000-0005"),
    ("Émile", "Dubois", CareerStage.postdoc, True, 3, None),
    ("Grace", "Kim", CareerStage.faculty, True, 4, "0000-0002-0000-0007"),
    ("Diego", "Martínez", CareerStage.grad, False, 2, None),
    ("Anna", "Kowalski", CareerStage.staff, True, 5, None),
    ("Ravi", "Patel", CareerStage.engineer, False, 1, None),
    ("Lucy", "Wright", CareerStage.undergrad, False, 0, None),
    ("Tomás", "Silva", CareerStage.postdoc, True, 5, "0000-0002-0000-0011"),
    ("Nina", "Volkova", CareerStage.faculty, True, 3, None),
]

# (title, conference idx, type, date, speaker idx or None, invited, status)
DEMO_EVENTS = [
    ("Future Colliders Workshop 2024", "Chicago, IL", date(2024, 9, 16), date(2024, 9, 20)),
    ("Accelerator Science Symposium 2025", "Berkeley, CA", date(2025, 3, 10), date(2025, 3, 14)),
    ("Particle Physics Frontiers 2025", "Knoxville, TN", date(2025, 10, 6), date(2025, 10, 10)),
    ("Community Summer Study 2026", "Boulder, CO", date(2026, 7, 20), date(2026, 7, 31)),
]

DEMO_TALKS = [
    ("Muon Collider Physics Overview", 0, "plenary", date(2024, 9, 17), 0, True, "given"),
    ("Cooling Channel Design Status", 0, "parallel", date(2024, 9, 18), 2, False, "given"),
    ("Detector Concepts for 10 TeV", 0, "parallel", date(2024, 9, 19), 7, False, "given"),
    ("High-Field Magnets for Muon Colliders", 1, "plenary", date(2025, 3, 11), 5, True, "given"),
    ("RF Systems for Rapid Acceleration", 1, "parallel", date(2025, 3, 12), 6, False, "given"),
    ("Beam-Induced Background Mitigation", 1, "parallel", date(2025, 3, 13), 2, True, "given"),
    ("Tracking in a High-Occupancy Environment", 2, "parallel", date(2025, 10, 7), 12, False, "given"),
    ("Calorimetry R&D Progress", 2, "parallel", date(2025, 10, 8), 3, False, "given"),
    ("US Muon Collider Program Status", 2, "plenary", date(2025, 10, 9), 1, True, "given"),
    ("Higgs Couplings at a Muon Collider", 3, "plenary", date(2026, 7, 21), 13, True, "assigned"),
    ("Neutrino Flux Opportunities", 3, "parallel", date(2026, 7, 23), 9, False, "assigned"),
    ("Machine-Detector Interface Design", 3, "parallel", date(2026, 7, 24), None, False, "nominations"),
]


@cli.command()
def seed_demo():
    """Populate the database with FICTIONAL demo data (for screenshots,
    demos, and development). Safe to run only on an empty/dev database."""
    from app.models import (
        AuthorList,
        Event,
        Nomination,
        Publication,
        PublicationPerson,
        PublicationPersonRole,
        PublicationStatus,
        PublicationType,
        Talk,
        TalkStatus,
        TalkType,
        User,
    )
    from app.services.author_list import build_snapshot

    with SessionLocal() as db:
        if db.execute(select(Person.id).limit(1)).first() is not None:
            typer.echo("Database already has people — refusing to seed demo data.")
            raise typer.Exit(1)

        insts = []
        for name, short, address in DEMO_INSTITUTIONS:
            inst = Institution(name=name, short_name=short, latex_address=address)
            db.add(inst)
            insts.append(inst)
        db.flush()

        people = []
        for i, (given, family, stage, voting, inst_idx, orcid) in enumerate(DEMO_PEOPLE):
            person = Person(
                given_name=given,
                family_name=family,
                email=f"{given}.{family}".lower()
                .replace(" ", "")
                .replace("'", "")
                .translate(str.maketrans("éíáóú", "eiaou"))
                + "@example.edu",
                orcid=orcid,
                career_stage=stage,
                status=MemberStatus.active,
                is_voting=voting,
                research_areas=(
                    "Accelerator Physics" if i % 2 else "Experimental Particle Physics"
                ),
                expertise="Muon collider R&D",
            )
            db.add(person)
            db.flush()
            # Spread joins over a few years so the growth chart has shape.
            joined = date(2023 + i % 3, 1 + i % 12, 1)
            db.add(
                Affiliation(
                    person_id=person.id,
                    institution_id=insts[inst_idx].id,
                    is_primary=True,
                    career_stage=stage,
                    start_date=joined,
                )
            )
            db.add(
                MembershipEvent(
                    person_id=person.id,
                    from_status=None,
                    to_status=MemberStatus.active.value,
                    effective_date=joined,
                )
            )
            if voting:
                db.add(AuthorPeriod(person_id=person.id, start_date=date(2024, 6, 1)))
            people.append(person)

        for name, slug in DEFAULT_WGS:
            if not db.execute(
                select(WorkingGroup).where(WorkingGroup.slug == slug)
            ).scalar_one_or_none():
                db.add(WorkingGroup(name=name, slug=slug))
        db.flush()

        # Leadership roles mirroring the organigram shapes (person idx, role,
        # detail, start, end).
        acc_wg = db.execute(
            select(WorkingGroup).where(WorkingGroup.slug == "accelerator")
        ).scalar_one()
        for p_idx, role, detail, start, end, wg_id in [
            (0, CollabRoleType.chair, None, date(2025, 1, 1), None, None),
            (1, CollabRoleType.vice_chair, None, date(2025, 1, 1), None, None),
            (3, CollabRoleType.representative, "Accelerator", date(2025, 1, 1), None, None),
            (7, CollabRoleType.representative, "Experimental", date(2025, 1, 1), None, None),
            (2, CollabRoleType.deputy_representative, "Experimental", date(2025, 6, 1), None, None),
            (13, CollabRoleType.coordinator, "Communications", date(2025, 1, 1), None, None),
            (9, CollabRoleType.area_lead, "Target", date(2025, 3, 1), None, None),
            (6, CollabRoleType.lsg_member, None, date(2024, 6, 1), None, None),
            (5, CollabRoleType.convener, None, date(2024, 9, 1), None, acc_wg.id),
            (12, CollabRoleType.chair, None, date(2023, 1, 1), date(2024, 12, 31), None),
        ]:
            db.add(
                CollabRole(
                    person_id=people[p_idx].id,
                    role=role,
                    detail=detail,
                    working_group_id=wg_id,
                    start_date=start,
                    end_date=end,
                )
            )
        # Administrative Institutional Contact (charter role, scoped to an
        # institution) — Fatima Hassan is staff at the second institution.
        db.add(
            CollabRole(
                person_id=people[5].id,
                role=CollabRoleType.admin_contact,
                institution_id=insts[1].id,
                start_date=date(2025, 1, 1),
            )
        )

        events = []
        for name, location, start, end in DEMO_EVENTS:
            ev = Event(name=name, location=location, start_date=start, end_date=end)
            db.add(ev)
            events.append(ev)
        db.flush()

        admin_user = db.execute(select(User).limit(1)).scalar_one_or_none()
        for title, ev_idx, ttype, tdate, sp_idx, invited, status in DEMO_TALKS:
            talk = Talk(
                title=title,
                event_id=events[ev_idx].id,
                talk_type=TalkType(ttype),
                date=tdate,
                speaker_person_id=people[sp_idx].id if sp_idx is not None else None,
                status=TalkStatus(status),
                is_invited=invited,
            )
            db.add(talk)
            db.flush()
            if status == "nominations":
                for cand in (6, 10):
                    db.add(
                        Nomination(
                            talk_id=talk.id,
                            person_id=people[cand].id,
                            nominated_by_user_id=admin_user.id if admin_user else None,
                        )
                    )

        pub = Publication(
            title="Detector Performance Studies for a 10 TeV Muon Collider",
            short_code="USMCC-WHIT-2026-001",
            pub_type=PublicationType.white_paper,
            status=PublicationStatus.collab_review,
            target_journal="arXiv",
            abstract="We present simulation studies of tracking, calorimetry, and "
            "beam-induced background rejection for a detector at a 10 TeV "
            "center-of-mass muon collider. (Fictional demo entry.)",
            author_cutoff_date=date(2026, 7, 1),
        )
        db.add(pub)
        db.flush()
        db.add(
            PublicationPerson(
                publication_id=pub.id,
                person_id=people[0].id,
                role=PublicationPersonRole.editor,
            )
        )
        snapshot = build_snapshot(db, date(2026, 7, 1))
        db.add(
            AuthorList(
                publication_id=pub.id,
                cutoff_date=date(2026, 7, 1),
                generated_by_user_id=admin_user.id if admin_user else None,
                snapshot=snapshot,
            )
        )
        db.commit()
        typer.echo(
            f"Seeded demo data: {len(DEMO_PEOPLE)} people, {len(DEMO_INSTITUTIONS)} "
            f"institutions, {len(DEMO_TALKS)} talks, 1 publication with author list."
        )


# --- Direct xlsx imports (match the USMCC Google-form exports) -----------------

ORCID_RE = re.compile(r"(\d{4}-\d{4}-\d{4}-\d{3}[\dX])")

POSITION_MAP = {
    "faculty": CareerStage.faculty,
    "postdoc": CareerStage.postdoc,
    "graduate student": CareerStage.grad,
    "undergraduate student": CareerStage.undergrad,
    "engineer": CareerStage.engineer,
    "accelerator engineer": CareerStage.engineer,
}


def _career_stage(position: str) -> CareerStage:
    p = (position or "").strip().lower()
    if p in POSITION_MAP:
        return POSITION_MAP[p]
    if "student" in p:
        return CareerStage.grad
    if "engineer" in p:
        return CareerStage.engineer
    if "scientist" in p or "researcher" in p or "physicist" in p or "staff" in p:
        return CareerStage.staff
    return CareerStage.other


def _clean_orcid(raw) -> str | None:
    if not raw:
        return None
    m = ORCID_RE.search(str(raw))
    return m.group(1) if m else None


def _split_expertise(raw: str | None) -> tuple[str | None, str | None]:
    """Split the form's Area(s) of Expertise cell into the standard research
    areas (the form's checkbox options, matched case-insensitively) and the
    leftover hand-entered values, which become free-form topics."""
    canonical = {a.lower(): a for a in RESEARCH_AREAS}
    areas: list[str] = []
    topics: list[str] = []
    for token in re.split(r"[,;]", raw or ""):
        token = token.strip()
        if not token:
            continue
        area = canonical.get(token.lower())
        if area is not None:
            if area not in areas:
                areas.append(area)
        elif token not in topics:
            topics.append(token)
    return ", ".join(areas) or None, ", ".join(topics) or None


def _percent_time(raw) -> int | None:
    """Map the form's percent-time answer to a single integer percent for
    ``usmcc_percent``. Range answers ('0-10%', '25-49%', ...) become the
    range midpoint; plain numbers are taken as-is; anything else ('Too
    unsure to estimate here', blank) becomes None."""
    if raw is None:
        return None
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        pct = round(raw)
        return pct if 0 <= pct <= 100 else None
    bounds = [int(n) for n in re.findall(r"\d+", str(raw)) if int(n) <= 100][:2]
    if not bounds:
        return None
    return round(sum(bounds) / len(bounds))


def _get_or_create_institution(db, name: str) -> Institution:
    name = name.strip()
    inst = db.execute(select(Institution).where(Institution.name == name)).scalar_one_or_none()
    if inst is None:
        inst = db.execute(
            select(Institution).where(Institution.short_name == name)
        ).scalar_one_or_none()
    if inst is None:
        inst = Institution(name=name)
        db.add(inst)
        db.flush()
    return inst


@cli.command()
def import_members_xlsx(
    xlsx_path: Path = typer.Argument(..., exists=True, readable=True),
    sheet: str = typer.Option("Members", help="Worksheet name"),
    authors_from_voting: bool = typer.Option(
        True, help="Open an author period for voting members (from registration date)"
    ),
    dry_run: bool = typer.Option(False, help="Parse and report, write nothing"),
):
    """Import the USMCC membership registration spreadsheet (Google-form
    export: Timestamp, voting question, First/Middle/Last Name, Primary
    Affiliation, additional affiliations, Email, ORCID, Position, ...,
    Area(s) of Expertise, percent of research time)."""
    import openpyxl

    from app.routers.people import _current_institution_is_us, _voting_eligible

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet]
    header = [str(c or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    def col_opt(label_start: str) -> int | None:
        for i, h in enumerate(header):
            if h.lower().startswith(label_start.lower()):
                return i
        return None

    def col(label_start: str) -> int:
        i = col_opt(label_start)
        if i is None:
            raise typer.BadParameter(f"Column starting with '{label_start}' not found in sheet")
        return i

    c_ts = col("Timestamp")
    c_voting = col("According to this definition")
    c_first, c_middle, c_last = col("First Name"), col("Middle Name"), col("Last Name")
    c_primary = col("Primary Affiliation")
    # The "Any additional affiliations" column is deliberately ignored:
    # secondary institutions are out of scope for now (issue #3, first pass).
    c_email, c_orcid, c_position = col("Email"), col("ORCID"), col("Position")
    c_expertise = col("Area(s) of Expertise")
    # The office-maintained "Simplified time" column is the cleaned-up version
    # of the form's percent-of-research-time question; use it when present.
    c_percent = col_opt("Simplified time")
    if c_percent is None:
        c_percent = col_opt("What percent of your research time")

    created = updated = skipped = 0
    with SessionLocal() as db:
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            email = str(row[c_email] or "").strip().lower()
            first = str(row[c_first] or "").strip()
            last = str(row[c_last] or "").strip()
            if not (email and first and last):
                if any(row):
                    typer.echo(f"row {i}: SKIP (missing name/email)")
                    skipped += 1
                continue
            middle = str(row[c_middle] or "").strip()
            given = f"{first} {middle}".strip() if middle else first
            # A blank voting answer is NOT a voting request — only an explicit
            # answer other than "non-voting" asks for the flag (issue #58).
            voting_answer = str(row[c_voting] or "").strip().lower()
            wants_voting = bool(voting_answer) and "non-voting" not in voting_answer
            orcid = _clean_orcid(row[c_orcid])
            stage = _career_stage(str(row[c_position] or ""))
            research_areas, expertise = _split_expertise(str(row[c_expertise] or ""))
            percent = _percent_time(row[c_percent]) if c_percent is not None else None
            ts = row[c_ts]
            start = ts.date() if hasattr(ts, "date") else date.today()

            person = db.execute(select(Person).where(Person.email == email)).scalar_one_or_none()
            if person is None and orcid:
                person = db.execute(
                    select(Person).where(Person.orcid == orcid)
                ).scalar_one_or_none()
            if person is None:
                person = Person(
                    given_name=given,
                    family_name=last,
                    preferred_name=first if middle else None,
                    email=email,
                    orcid=orcid,
                    career_stage=stage,
                    status=MemberStatus.active,
                    # Stamped below, once the affiliation is in place and the
                    # charter eligibility rules can be checked.
                    is_voting=False,
                    research_areas=research_areas,
                    expertise=expertise,
                    usmcc_percent=percent,
                )
                db.add(person)
                db.flush()
                created += 1
            else:
                person.given_name, person.family_name = given, last
                person.orcid = orcid or person.orcid
                person.career_stage = stage
                person.research_areas = research_areas or person.research_areas
                person.expertise = expertise or person.expertise
                if percent is not None:
                    person.usmcc_percent = percent
                updated += 1
            _ensure_activation_event(db, person, start)

            primary_name = str(row[c_primary] or "").strip()
            if primary_name:
                inst = _get_or_create_institution(db, primary_name)
                _set_primary_affiliation(db, f"row {i}", person, inst, stage, start)

            # Stamp the voting flag only now, with the affiliation in place,
            # applying the same charter rules the API enforces: active,
            # non-student, currently at a US institution. A blank answer on a
            # re-import keeps the stored flag (still subject to the rules).
            db.flush()  # session is autoflush=False; expose the new affiliation
            voting = wants_voting if voting_answer else person.is_voting
            if voting and not (
                _voting_eligible(person.status, person.career_stage)
                and _current_institution_is_us(db, person.id)
            ):
                typer.echo(
                    f"row {i}: voting flag dropped (requires an active, "
                    "non-student member at a US institution)"
                )
                voting = False
            person.is_voting = voting

            if authors_from_voting and voting:
                has_period = db.execute(
                    select(AuthorPeriod).where(
                        AuthorPeriod.person_id == person.id,
                        AuthorPeriod.end_date.is_(None),
                    )
                ).scalar_one_or_none()
                if has_period is None:
                    db.add(AuthorPeriod(person_id=person.id, start_date=start))

        if dry_run:
            db.rollback()
            typer.echo(f"DRY RUN — would create {created}, update {updated}, skip {skipped}")
        else:
            db.commit()
            typer.echo(f"Imported: {created} created, {updated} updated, {skipped} skipped")


@cli.command()
def import_talks_xlsx(
    xlsx_path: Path = typer.Argument(..., exists=True, readable=True),
    sheet: str = typer.Option("Assigned Talks", help="Worksheet name"),
    dry_run: bool = typer.Option(False, help="Parse and report, write nothing"),
):
    """Import the Conferences_and_Speakers spreadsheet (columns: Date,
    Conference, Topic, Name, Plenary/Parallel, Invited/Contributed, URL,
    Notes). Creates events by conference name and matches speakers to
    people by full name."""
    import openpyxl

    from app.models import Event, Talk, TalkStatus, TalkType

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet]

    with SessionLocal() as db:
        people = db.execute(select(Person)).scalars().all()
        by_name: dict[str, Person] = {}
        for p in people:
            by_name[f"{p.given_name} {p.family_name}".lower()] = p
            first = p.given_name.split()[0] if p.given_name else ""
            by_name.setdefault(f"{first} {p.family_name}".lower(), p)
            if p.preferred_name:
                by_name.setdefault(f"{p.preferred_name} {p.family_name}".lower(), p)

        events: dict[str, Event] = {}
        created = unmatched = skipped = 0
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            when, conf, topic, name = row[0], row[1], row[2], row[3]
            if not (topic or name):
                continue
            title = str(topic or "(untitled)").strip()
            conf_name = str(conf or "").strip()
            talk_date = when.date() if hasattr(when, "date") else None
            plenary = "plenary" in str(row[4] or "").lower()
            invited = "invited" in str(row[5] or "").lower()
            url = str(row[6] or "").strip()
            notes = str(row[7] or "").strip()

            event = None
            if conf_name:
                if conf_name not in events:
                    existing = db.execute(
                        select(Event).where(Event.name == conf_name)
                    ).scalar_one_or_none()
                    events[conf_name] = existing or Event(name=conf_name)
                    if existing is None:
                        db.add(events[conf_name])
                        db.flush()
                event = events[conf_name]

            speaker = by_name.get(str(name or "").strip().lower())
            extra_notes = []
            if url:
                extra_notes.append(url)
            if notes:
                extra_notes.append(notes)
            if speaker is None and name:
                unmatched += 1
                extra_notes.append(f"speaker (unmatched): {name}")

            # Skip exact duplicates on re-import.
            dupe = db.execute(
                select(Talk).where(
                    Talk.title == title,
                    Talk.event_id == (event.id if event else None),
                    Talk.date == talk_date,
                )
            ).scalar_one_or_none()
            if dupe:
                skipped += 1
                continue

            db.add(
                Talk(
                    title=title,
                    event_id=event.id if event else None,
                    talk_type=TalkType.plenary if plenary else TalkType.parallel,
                    date=talk_date,
                    speaker_person_id=speaker.id if speaker else None,
                    status=TalkStatus.given if talk_date and talk_date <= date.today() else TalkStatus.assigned,
                    is_invited=invited,
                    notes="\n".join(extra_notes) or None,
                )
            )
            created += 1

        if dry_run:
            db.rollback()
            typer.echo(
                f"DRY RUN — would create {created} talks "
                f"({unmatched} unmatched speakers, {skipped} duplicates skipped)"
            )
        else:
            db.commit()
            typer.echo(
                f"Imported {created} talks ({unmatched} speakers unmatched — "
                f"kept in notes; {skipped} duplicates skipped)"
            )


# --- Photo imports ---------------------------------------------------------------

DRIVE_ID_RE = re.compile(r"(?:id=|/d/)([-\w]{20,})")
IMAGE_EXTS = {".jpg": ".jpg", ".jpeg": ".jpg", ".png": ".png", ".webp": ".webp", ".gif": ".gif"}
CONTENT_EXTS = {
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
    "image/gif": ".gif",
}


def _photos_dir() -> Path:
    from app.config import get_settings

    d = Path(get_settings().photos_dir)
    d.mkdir(parents=True, exist_ok=True)
    return d


def _save_photo(db, person: Person, content: bytes, ext: str) -> None:
    from datetime import datetime, timezone

    name = f"{person.id}-{int(datetime.now(timezone.utc).timestamp())}{ext}"
    (_photos_dir() / name).write_bytes(content)
    old = person.photo_file
    person.photo_file = name
    db.flush()
    if old and old != name and (_photos_dir() / old).is_file():
        (_photos_dir() / old).unlink()


@cli.command()
def import_photos_xlsx(
    xlsx_path: Path = typer.Argument(..., exists=True, readable=True),
    sheet: str = typer.Option("Members", help="Worksheet name"),
    overwrite: bool = typer.Option(False, help="Replace photos that already exist"),
):
    """Download the member photos linked in the registration spreadsheet's
    'Photo link' column (Google Drive links).

    NOTE: Google-Form uploads are often restricted to the form owner. Links
    that are not shared 'anyone with the link' will fail here — download
    those from your Drive and use import-photos-dir instead."""
    import httpx
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet]
    header = [str(c or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    try:
        c_email = next(i for i, h in enumerate(header) if h.lower().startswith("email"))
        c_photo = next(i for i, h in enumerate(header) if h.lower().startswith("photo"))
    except StopIteration:
        raise typer.BadParameter("Email / Photo link columns not found")

    ok = skipped = failed = 0
    failures: list[str] = []
    with SessionLocal() as db, httpx.Client(follow_redirects=True, timeout=30) as client:
        for row in ws.iter_rows(min_row=2, values_only=True):
            email = str(row[c_email] or "").strip().lower()
            link = str(row[c_photo] or "").strip()
            if not (email and link):
                continue
            person = db.execute(select(Person).where(Person.email == email)).scalar_one_or_none()
            if person is None:
                failures.append(f"{email}: not in database")
                failed += 1
                continue
            if person.photo_file and not overwrite:
                skipped += 1
                continue
            m = DRIVE_ID_RE.search(link)
            url = (
                f"https://drive.google.com/uc?export=download&id={m.group(1)}"
                if m
                else link
            )
            try:
                resp = client.get(url)
                ctype = resp.headers.get("content-type", "").split(";")[0].strip()
                if resp.status_code != 200 or ctype not in CONTENT_EXTS:
                    # Large-file interstitial: retry via usercontent endpoint.
                    if m and "text/html" in ctype:
                        resp = client.get(
                            "https://drive.usercontent.google.com/download",
                            params={"id": m.group(1), "export": "download", "confirm": "t"},
                        )
                        ctype = resp.headers.get("content-type", "").split(";")[0].strip()
                if resp.status_code != 200 or ctype not in CONTENT_EXTS:
                    failures.append(
                        f"{person.given_name} {person.family_name}: got {resp.status_code} "
                        f"{ctype or 'unknown type'} (probably not shared publicly)"
                    )
                    failed += 1
                    continue
                _save_photo(db, person, resp.content, CONTENT_EXTS[ctype])
                ok += 1
            except httpx.HTTPError as exc:
                failures.append(f"{person.given_name} {person.family_name}: {exc}")
                failed += 1
        db.commit()

    typer.echo(f"Photos: {ok} downloaded, {skipped} already present, {failed} failed")
    if failures:
        typer.echo("\nFailed (fix sharing, or bulk-download from Drive and run import-photos-dir):")
        for f in failures:
            typer.echo(f"  - {f}")


@cli.command()
def import_photos_dir(
    dir_path: Path = typer.Argument(..., exists=True, file_okay=False),
    overwrite: bool = typer.Option(False, help="Replace photos that already exist"),
):
    """Import photos from a directory, matching people by name in the file
    name (Google-Form uploads are named like 'IMG_1234 - Jane Doe.jpg').
    Unmatched files are listed at the end."""
    ok = skipped = 0
    unmatched: list[str] = []
    with SessionLocal() as db:
        people = db.execute(select(Person)).scalars().all()
        # All name variants, longest first so "Mary Jane Smith" wins over "Jane Smith".
        variants: list[tuple[str, Person]] = []
        for p in people:
            names = {
                f"{p.given_name} {p.family_name}",
                f"{p.given_name.split()[0]} {p.family_name}" if p.given_name else "",
                f"{p.preferred_name} {p.family_name}" if p.preferred_name else "",
            }
            variants.extend((n.lower(), p) for n in names if n)
        variants.sort(key=lambda v: -len(v[0]))

        for path in sorted(dir_path.iterdir()):
            ext = IMAGE_EXTS.get(path.suffix.lower())
            if not path.is_file() or ext is None:
                continue
            stem = path.stem.lower()
            person = next((p for n, p in variants if n in stem), None)
            if person is None:
                unmatched.append(path.name)
                continue
            if person.photo_file and not overwrite:
                skipped += 1
                continue
            _save_photo(db, person, path.read_bytes(), ext)
            ok += 1
        db.commit()

    typer.echo(f"Photos: {ok} imported, {skipped} already present, {len(unmatched)} unmatched")
    for name in unmatched:
        typer.echo(f"  unmatched: {name}")


if __name__ == "__main__":
    cli()
